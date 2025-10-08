from django.shortcuts import render,redirect
from dashboard.models import GM, RSM , UserProfile
from restaurant.models import Restaurant,Restaurant_Target,Restaurant_Crate_Sales,CoolerSettings,PosmSettings,Restaurant_Cooler,Restaurant_POSM,Manager
from dashboard.models import Bottler
from consumers.models import Consumer
from django.contrib.auth.models import User
import json
from django.http import JsonResponse
from rest_framework_simplejwt.tokens import RefreshToken
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate , login
from django.contrib.auth.decorators import login_required
from django.forms.models import model_to_dict
from django.db.models import Sum ,Q,Avg
from rest_framework_simplejwt.tokens import AccessToken
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.core.paginator import Paginator
from django.db.models import Sum, Avg
from django.db.models.functions import ExtractMonth,ExtractDay
import calendar
from django.utils.dateparse import parse_date
from consumers.models import Consumer
from django.db.models import Count
from django.contrib.auth import logout
from datetime import datetime,timedelta
from .decorators import superuser_required,developer_login_required
import re
from django.utils.decorators import method_decorator


def login_view(request):
    if request.user.is_authenticated:
        return redirect("/dashboard/index/")

    return render(request, "index.html")   


def login_api(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            username = data.get("username")
            password = data.get("password")

            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                refresh = RefreshToken.for_user(user)
                request.session['user_id'] = user.id

                userprofile = UserProfile.objects.get(user=user.id)
                if userprofile.role == "Developer":
                    return JsonResponse({
                    "success": True,
                    "message": "Developer Login successful",
                    "access": str(refresh.access_token),
                    "redirect_url": "/dashboard/developer-page/",
                    "refresh": str(refresh),
                })
                else:    
                    return JsonResponse({
                        "success": True,
                        "message": "Login successful",
                        "access": str(refresh.access_token),
                        "redirect_url": "/dashboard/index/",
                        "refresh": str(refresh),
                    })
            else:
                return JsonResponse({"success": False, "message": "Invalid credentials"})
        except json.JSONDecodeError:
            return JsonResponse({"success": False, "message": "Invalid JSON format"})

    return JsonResponse({"success": False, "message": "Only POST allowed"})


@developer_login_required
def developer_page(request):
    user_id = request.session.get("user_id")
    user = User.objects.get(id=user_id)
    username = user.username

    restaurant_coolers = Restaurant_Cooler.objects.filter(
    is_checked=False
).order_by("-created_at")

    cooler_list = []

    for cooler in restaurant_coolers:
        compliance = 0
        if cooler.raw_data and "/" in cooler.raw_data:
            try:
                num, denom = cooler.raw_data.split("/")
                num = int(num)
                denom = int(denom)
                compliance = round((num / denom) * 100, 2) if denom else 0
            except ValueError:
                compliance = 0

        # determine status text
        print(cooler.is_checked)
        if not cooler.is_checked:
            status = "Requested"
        elif hasattr(cooler, "is_checked") and cooler.is_checked:
            status = "Approved"
        else:
            status = "-"

        cooler_list.append({
            "id": cooler.id,
            "cooler_type": cooler.cooler_type,
            "raw_data": cooler.raw_data,
            "compliance": compliance,
            "status": status,
            "image": cooler.image,
            "created_at": cooler.created_at.strftime("%Y-%m-%d %H:%M"),  # cleaner format
        })



    return render(request, "developer_page.html",
                   {
                       "username":username,
        "cooler_list":cooler_list
        },
)





@method_decorator(developer_login_required, name='dispatch')
class CoolerDefaultAPIView(APIView):
    """
    GET:  return {"default_is_checked": bool}
    POST: accept {"default_is_checked": true|false} to update the setting
    """
    def get(self, request):
        settings, created = CoolerSettings.objects.get_or_create(id=1)
        return Response({"default_is_checked": settings.default_is_checked}, status=status.HTTP_200_OK)

    def post(self, request):
        # Accept boolean or string "true"/"false"
        default_value = request.data.get("default_is_checked", True)
        # Normalize common string values
        if isinstance(default_value, str):
            default_value = default_value.lower() in ("1", "true", "yes", "on")

        settings, created = CoolerSettings.objects.get_or_create(id=1)
        settings.default_is_checked = bool(default_value)
        settings.save()

        return Response({"default_is_checked": settings.default_is_checked}, status=status.HTTP_200_OK)
    


@developer_login_required
def developer_page_posm(request):
    user_id = request.session.get("user_id")
    user = User.objects.get(id=user_id)
    username = user.username

    restaurant_posm = Restaurant_POSM.objects.filter(
    is_checked=False
).order_by("-created_at")

    posm_list = []

    for posm in restaurant_posm:
        compliance = 0
        if posm.raw_data and "/" in posm.raw_data:
            try:
                num, denom = posm.raw_data.split("/")
                num = int(num)
                denom = int(denom)
                compliance = round((num / denom) * 100, 2) if denom else 0
            except ValueError:
                compliance = 0

        # determine status text
        print(posm.is_checked)
        if not posm.is_checked:
            status = "Requested"
        elif hasattr(posm, "is_checked") and posm.is_checked:
            status = "Approved"
        else:
            status = "-"

        posm_list.append({
            "id": posm.id,
            "posm_type": posm.posm_type,
            "raw_data": posm.raw_data,
            "compliance": compliance,
            "status": status,
            "image": posm.image,
            "created_at": posm.created_at.strftime("%Y-%m-%d %H:%M"),  # cleaner format
        })



    return render(request, "developer-page-posm.html",
                   {
                       "username":username,
        "posm_list":posm_list
        },
)





@method_decorator(developer_login_required, name='dispatch')
class PosmDefaultAPIView(APIView):
    """
    GET:  return {"default_is_checked": bool}
    POST: accept {"default_is_checked": true|false} to update the setting
    """
    def get(self, request):
        settings, created = PosmSettings.objects.get_or_create(id=1)
        return Response({"default_is_checked": settings.default_is_checked}, status=status.HTTP_200_OK)

    def post(self, request):
        # Accept boolean or string "true"/"false"
        default_value = request.data.get("default_is_checked", True)
        # Normalize common string values
        if isinstance(default_value, str):
            default_value = default_value.lower() in ("1", "true", "yes", "on")

        settings, created = PosmSettings.objects.get_or_create(id=1)
        settings.default_is_checked = bool(default_value)
        settings.save()

        return Response({"default_is_checked": settings.default_is_checked}, status=status.HTTP_200_OK)
    

@login_required(login_url='/dashboard/login/')
def dashboard(request):
    user_id = request.session.get("user_id")
    user = User.objects.get(id=user_id)
    username = user.username
    # --- Get Filters from request ---
    start_date = request.GET.get("startDate")
    end_date = request.GET.get("endDate") 
    bottler_id = request.GET.get("bottler")
    gm_id = request.GET.get("gm")
    rsm_id = request.GET.get("rsm")
    restaurant_id = request.GET.get("restaurant")



    # Convert dates
    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)


    # --- Build filters ---
    totalbottlerfilter = {}
    bottlerbasefilter = {}
    bottlerbaserestaurantfilter = {}
    totaltargetsfilter = {}
    gmbasefilter = {}
    rsmbasefilter = {}
    restaurantbasefilter = {}
    bottlergmbase = {}
    bottlerrsmbase = {}
    bottlerrestaurantbase = {}
    gmbaserestaurantfilter = {}
    rsmbaserestaurantfilter = {}
    restaurantbaserestaurantfilter = {}


    end_date_plus_one = 0
    if end_date:    
        end_date_plus_one = end_date + timedelta(days=1)

    if start_date and end_date:
        if bottler_id and not gm_id:
            print("bottler condition called line 99")
            totalbottlerfilter["id__in"] = bottler_id
            totaltargetsfilter["restaurant__bottler"] = bottler_id
            bottlerbasefilter["restaurant__bottler"] = bottler_id
            if gm_id and not rsm_id:
                gm = GM.objects.get(id=gm_id)
                gm_restaurants = gm.restaurants.all()
                bottlergmbase["id"] = gm.bottler_id
                gmbasefilter["restaurant__in"] = gm_restaurants
                gmbasefilter["created_at__gte"] = start_date
                gmbasefilter["created_at__lt"] = end_date_plus_one
                gmbasefilter["created_at__year__gte"] = start_date.year
                gmbasefilter["created_at__month__gte"] = start_date.month
                gmbasefilter["created_at__year__lte"] = end_date_plus_one.year
                gmbasefilter["created_at__month__lte"] = end_date_plus_one.month
                if rsm_id and not restaurant_id:
                    rsm = RSM.objects.get(id=rsm_id)
                    bottlerrsmbase["id"]=rsm.bottler.id
                    rsm_restaurants = rsm.restaurants.all()
                    rsmbasefilter["restaurant__in"] = rsm_restaurants
                    rsmbasefilter["created_at__gte"] = start_date
                    rsmbasefilter["created_at__lt"] = end_date_plus_one
                    rsmbasefilter["created_at__year__gte"] = start_date.year
                    rsmbasefilter["created_at__month__gte"] = start_date.month
                    rsmbasefilter["created_at__year__lte"] = end_date_plus_one.year
                    rsmbasefilter["created_at__month__lte"] = end_date_plus_one.month
                    if restaurant_id:
                        restaurant = Restaurant.objects.filter(id=restaurant_id)
                        bottlerrestaurantbase["id"]=restaurant.bottler.id
                        restaurantbasefilter["created_at__gte"] = start_date
                        restaurantbasefilter["created_at__lt"] = end_date_plus_one
                        restaurantbasefilter["created_at__year__gte"] = start_date.year
                        restaurantbasefilter["created_at__month__gte"] = start_date.month
                        restaurantbasefilter["created_at__year__lte"] = end_date_plus_one.year
                        restaurantbasefilter["created_at__month__lte"] = end_date_plus_one.month

        bottlerbasefilter["created_at__gte"] = start_date
        bottlerbasefilter["created_at__lt"] = end_date_plus_one
        totaltargetsfilter["created_at__year__gte"] = start_date.year
        totaltargetsfilter["created_at__month__gte"] = start_date.month
        totaltargetsfilter["created_at__year__lte"] = end_date_plus_one.year
        totaltargetsfilter["created_at__month__lte"] = end_date_plus_one.month


    if bottler_id and not gm_id:
        print("bottler condition called line 123")
        totalbottlerfilter["id__in"] = bottler_id
        bottlerbasefilter["restaurant__bottler"] = bottler_id
        bottlerbaserestaurantfilter["bottler_id"] = bottler_id
        totaltargetsfilter["restaurant__bottler"] = bottler_id
    
    if gm_id and not rsm_id:
        gm = GM.objects.get(id=gm_id)
        gm_restaurants = gm.restaurants.all()
        gmbasefilter["restaurant__in"] = gm_restaurants
        gmbaserestaurantfilter["id__in"] = gm_restaurants
        bottlergmbase["id"]=gm.bottler_id

    if rsm_id and not gm_id and not restaurant_id:
        rsm = RSM.objects.get(id=rsm_id)
        rsm_restaurants = rsm.restaurants.all()
        rsmbasefilter["restaurant__in"] = rsm_restaurants
        bottlerrsmbase["id"] = rsm.bottler.id
        rsmbaserestaurantfilter["id__in"] = rsm_restaurants
    if restaurant_id:
        restaurant = Restaurant.objects.get(id=restaurant_id)
        restaurantbasefilter["restaurant"] = restaurant
        restaurantbaserestaurantfilter["id"] = restaurant.id
    # Initial values
    bottle_crates_total = 0
    bottle_crate_target_achieved = 0
    cooler_compliance = 0
    posm_compliance = 0
    total_compliance_target = 0
    total_compliance_achieved = 0
    total_bottler = 0
    total_restaurant = 0
    consumer_total = 0
    consumer_target_achieved = 0
    sales_achievement = 0
    consumer_achievement = 0
    overall_target_achieved = 0

    if user.is_superuser:
        bottlers = Bottler.objects.all()
        restaurants = Restaurant.objects.all()
        gms = GM.objects.all()
        rsms = RSM.objects.all()

        # Cases Purchased
        
        bottle_crates_total = Restaurant_Target.objects.filter(
            target_type="Cases Purchased",**totaltargetsfilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter
        ).aggregate(total=Sum("target_value"))["total"] or 0

        bottle_crate_target_achieved = Restaurant_Crate_Sales.objects.filter(
            is_approved=True,**bottlerbasefilter,**gmbasefilter,**rsmbasefilter , **restaurantbasefilter 
        ).aggregate(total=Sum("crate_quantity"))["total"] or 0

        # Compliance (Cooler)
        coolers = Restaurant_Cooler.objects.filter(is_checked=True,**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)
        cooler_num_sum, cooler_den_sum = 0, 0
        for cooler in coolers:
            if cooler.raw_data and "/" in cooler.raw_data:
                try:
                    num, denom = map(int, cooler.raw_data.split("/"))
                    cooler_num_sum += num
                    cooler_den_sum += denom
                except ValueError:
                    continue
        cooler_compliance = (
            round((cooler_num_sum / cooler_den_sum) * 100, 2)
            if cooler_den_sum else 0
        )

        # Compliance (POSM)
        posms = Restaurant_POSM.objects.filter(is_checked=True,**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)
        posm_num_sum, posm_den_sum = 0, 0
        for posm in posms:
            if posm.raw_data and "/" in posm.raw_data:
                try:
                    num, denom = map(int, posm.raw_data.split("/"))
                    posm_num_sum += num
                    posm_den_sum += denom
                except ValueError:
                    continue
        posm_compliance = (
            round((posm_num_sum / posm_den_sum) * 100, 2)
            if posm_den_sum else 0
        )

        total_compliance_target = (
            Restaurant_Target.objects.filter(target_type="Compliance",**totaltargetsfilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)
            .aggregate(total=Avg("target_value"))["total"] or 0
        )
        total_compliance_achieved = (cooler_compliance * (2 / 3)) + (posm_compliance * (1 / 3))

        print(total_compliance_achieved)

        # Restaurant / Bottler count
        total_restaurant = Restaurant.objects.filter(**bottlerbaserestaurantfilter,**gmbaserestaurantfilter,**rsmbaserestaurantfilter,**restaurantbaserestaurantfilter).count()
        total_bottler = Bottler.objects.filter(**totalbottlerfilter,**bottlergmbase,**bottlerrsmbase).count()

        # Consumer
        consumer_total = Restaurant_Target.objects.filter(
            target_type="Consumer",**totaltargetsfilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter 
        ).aggregate(total=Sum("target_value"))["total"] or 0
        consumer_target_achieved = Consumer.objects.filter(is_checked=True,**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter).count()

        # Overall target %
        if bottle_crates_total > 0:
            sales_achievement = round((bottle_crate_target_achieved / bottle_crates_total) * 100, 2) if bottle_crates_total else 0
        if consumer_total > 0:
            consumer_achievement = round((consumer_target_achieved / consumer_total) * 100, 2) if consumer_total else 0
        

        overall_target_achieved = round(
            (cooler_compliance * 0.2) + (posm_compliance * 0.1) +
            (sales_achievement * 0.5) + (consumer_achievement * 0.2), 2
        )

        # Sales Month Graph Data
        volumes = (
            Restaurant_Target.objects.filter(target_type="Cases Purchased", target_status=True,**totaltargetsfilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)
            .annotate(month=ExtractMonth("created_at"))
            .values("month")
            .annotate(total_target=Sum("target_value"))
            .order_by("month")
        )
        targets = (
            Restaurant_Crate_Sales.objects.filter(is_approved=True,**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)
            .annotate(month=ExtractMonth("created_at"))
            .values("month")
            .annotate(crate_quantity=Sum("crate_quantity"))
            .order_by("month")
        )
        month_labels = [calendar.month_abbr[i] for i in range(1, 13)]
        sales_target = [0] * 12
        sales_volume = [0] * 12
        for t in targets:
            sales_target[t["month"] - 1] = round(t["crate_quantity"], 2)
        for v in volumes:
            sales_volume[v["month"] - 1] = v["total_target"]

        
        # Compliance Graph
        compliance_volumes = (
            Restaurant_Target.objects.filter(target_type="Compliance", target_status=True,**totaltargetsfilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)
            .annotate(month=ExtractMonth("created_at"))
            .values("month")
            .annotate(total_target=Avg("target_value"))
            .order_by("month")
        )
        compliance_month_labels = [calendar.month_abbr[i] for i in range(1, 13)]
        compliance_target = [0] * 12
        compliance_volume = [0] * 12
        for v in compliance_volumes:
            compliance_target[v["month"] - 1] = round(v["total_target"], 2)
        for month in range(1, 13):
            coolers = Restaurant_Cooler.objects.filter(is_checked=True, created_at__month=month,**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)
            cooler_num_sum, cooler_den_sum = 0, 0
            for cooler in coolers:
                if cooler.raw_data and "/" in cooler.raw_data:
                    try:
                        num, denom = map(int, cooler.raw_data.split("/"))
                        cooler_num_sum += num
                        cooler_den_sum += denom
                    except ValueError:
                        continue
            cooler_compliance = round((cooler_num_sum / cooler_den_sum) * 100, 2) if cooler_den_sum else 0
            posms = Restaurant_POSM.objects.filter(is_checked=True, created_at__month=month,**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)
            posm_num_sum, posm_den_sum = 0, 0
            for posm in posms:
                if posm.raw_data and "/" in posm.raw_data:
                    try:
                        num, denom = map(int, posm.raw_data.split("/"))
                        posm_num_sum += num
                        posm_den_sum += denom
                    except ValueError:
                        continue
            posm_compliance = round((posm_num_sum / posm_den_sum) * 100, 2) if posm_den_sum else 0
            total_compliance = (cooler_compliance * (2 / 3)) + (posm_compliance * (1 / 3))
            compliance_volume[month - 1] = round(total_compliance, 2)

        # Consumer Graph
        consumer_volumes = (
            Restaurant_Target.objects.filter(target_type="Consumer", target_status=True,**totaltargetsfilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)
            .annotate(month=ExtractMonth("created_at"))
            .values("month")
            .annotate(total_target=Sum("target_value"))
            .order_by("month")
        )
        consumer_targets = (
            Consumer.objects.filter(is_checked=True,**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)
            .annotate(month=ExtractMonth("created_at"))
            .values("month")
            .annotate(total_count=Count("id"))
            .order_by("month")
        )
        consumer_month_labels = [calendar.month_abbr[i] for i in range(1, 13)]
        consumer_target = [0] * 12
        consumer_volume = [0] * 12
        for t in consumer_targets:
            if t["month"]:
                consumer_target[t["month"] - 1] = t["total_count"]
        for v in consumer_volumes:
            if v["month"]:
                consumer_volume[v["month"] - 1] = v["total_target"] or 0



    else:
        try:

            gm = GM.objects.get(user_id=user_id)
            gm_restaurants = gm.restaurants.all()

            bottlers = Bottler.objects.filter(id=gm.bottler.id)
            restaurants = Restaurant.objects.filter(id__in=gm_restaurants.values_list("id", flat=True))
            gms = GM.objects.filter(restaurants__in=gm_restaurants).distinct()
            rsms = RSM.objects.filter(restaurants__in=gm_restaurants).distinct()

            bottle_crates_total = Restaurant_Target.objects.filter(
                restaurant__in=gm_restaurants,
                **totaltargetsfilter,
                **gmbasefilter,
                **rsmbasefilter,**restaurantbasefilter,
                target_type="Cases Purchased"
            ).aggregate(total=Sum("target_value"))["total"] or 0

            bottle_crate_target_achieved = Restaurant_Crate_Sales.objects.filter(
                restaurant__in = gm_restaurants,
                **bottlerbasefilter,
                **gmbasefilter,
                **rsmbasefilter,**restaurantbasefilter,
                is_approved = True
            ).aggregate(total=Sum("crate_quantity"))["total"] or 0        


            coolers = Restaurant_Cooler.objects.filter(restaurant__in=gm_restaurants,is_checked=True,**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)

            cooler_numerator_sum = 0
            cooler_denominator_sum = 0

            for cooler in coolers:
                if cooler.raw_data and "/" in cooler.raw_data:
                    num, denom = cooler.raw_data.split("/")
                    try:
                        cooler_numerator_sum += int(num)
                        cooler_denominator_sum += int(denom)
                    except ValueError:
                        continue

            cooler_compliance = round((cooler_numerator_sum / cooler_denominator_sum) * 100, 2) if cooler_denominator_sum else 0

            posms = Restaurant_POSM.objects.filter(restaurant__in=gm_restaurants,is_checked=True,**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)

            posm_numerator_sum = 0
            posm_denominator_sum = 0

            for posm in posms:
                if posm.raw_data and "/" in posm.raw_data:
                    num, denom = posm.raw_data.split("/")
                    try:
                        posm_numerator_sum += int(num)
                        posm_denominator_sum += int(denom)
                    except ValueError:
                        continue

            posm_compliance = round((posm_numerator_sum / posm_denominator_sum) * 100, 2) if posm_denominator_sum else 0
    

            total_compliance_target = Restaurant_Target.objects.filter(
                restaurant__in=gm_restaurants,
                **totaltargetsfilter,
                **gmbasefilter,
                **rsmbasefilter,**restaurantbasefilter,
                target_type="Compliance"
            ).aggregate(total=Avg("target_value"))["total"] or 0

            total_compliance_achieved = (cooler_compliance*(2/3))+(posm_compliance*(1/3))

            consumer_total = Restaurant_Target.objects.filter(
                restaurant__in=gm_restaurants,
                **totaltargetsfilter,
                **gmbasefilter,
                **rsmbasefilter,**restaurantbasefilter,
                target_type="Consumer"
            ).aggregate(total=Sum("target_value"))["total"] or 0

            consumer_target_achieved = Consumer.objects.filter(
                    restaurant__in=gm_restaurants,
                    **bottlerbasefilter,
                    **gmbasefilter,
                    **rsmbasefilter,**restaurantbasefilter,
                    is_checked=True
                ).count()

            total_restaurant = gm_restaurants.count()
            total_bottler = Bottler.objects.filter(id=gm.bottler.id,**totalbottlerfilter,**bottlerrsmbase,**bottlerrestaurantbase).count

            if bottle_crates_total > 0:
                sales_achievement = round((bottle_crate_target_achieved / bottle_crates_total)*100,2)
            if consumer_total > 0:
                consumer_achievement = round((consumer_target_achieved/consumer_total)*100,2)
        
            overall_target_achieved = round((cooler_compliance*0.2)+(posm_compliance*0.1)+(sales_achievement*0.5)+(consumer_achievement*0.2),2)

# Sales Graph
            volumes = (
            Restaurant_Target.objects.filter(**totaltargetsfilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter,restaurant__in=gm_restaurants,target_type="Cases Purchased",target_status=True)
            .annotate(month=ExtractMonth("created_at"))
            .values("month")
            .annotate(total_target=Sum("target_value"))
            .order_by("month")
        )
            
            targets = (
                Restaurant_Crate_Sales.objects.filter(**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter,restaurant__in=gm_restaurants,is_approved=True)
                .annotate(month=ExtractMonth("created_at"))
                .values("month")
                .annotate(crate_quantity=Sum("crate_quantity"))
                .order_by("month")
            )


            month_labels = [calendar.month_abbr[i] for i in range(1, 13)]
            sales_target = [0] * 12
            sales_volume = [0] * 12

            for t in targets:
                sales_target[t["month"] - 1] = round(t["crate_quantity"], 2)

            for v in volumes:
                sales_volume[v["month"] - 1] = v["total_target"]


# compliance graph data
            compliance_volumes = (
                    Restaurant_Target.objects.filter(**totaltargetsfilter,**gmbasefilter,restaurant__in=gm_restaurants,**rsmbasefilter,**restaurantbasefilter,target_type="Compliance", target_status=True)
                    .annotate(month=ExtractMonth("created_at"))
                    .values("month")
                    .annotate(total_target=Avg("target_value"))
                    .order_by("month")
                )

            compliance_month_labels = [calendar.month_abbr[i] for i in range(1, 13)]
            compliance_target = [0] * 12  
            compliance_volume = [0] * 12  

            for v in compliance_volumes:
                compliance_target[v["month"] - 1] = round(v["total_target"], 2)

            for month in range(1, 13):
                # --- Coolers ---
                coolers = Restaurant_Cooler.objects.filter(
                    restaurant__in=gm_restaurants,
                    **bottlerbasefilter,
                    **gmbasefilter,
                    **rsmbasefilter,**restaurantbasefilter,
                    is_checked=True,
                    created_at__month=month
                )

                cooler_num_sum, cooler_den_sum = 0, 0
                for cooler in coolers:
                    if cooler.raw_data and "/" in cooler.raw_data:
                        try:
                            num, denom = map(int, cooler.raw_data.split("/"))
                            cooler_num_sum += num
                            cooler_den_sum += denom
                        except ValueError:
                            continue

                cooler_compliance = (
                    round((cooler_num_sum / cooler_den_sum) * 100, 2)
                    if cooler_den_sum else 0
                )

                # --- POSMs ---
                posms = Restaurant_POSM.objects.filter(
                    restaurant__in=gm_restaurants,
                    **bottlerbasefilter,
                    **gmbasefilter,
                    **rsmbasefilter,**restaurantbasefilter,
                    is_checked=True,
                    created_at__month=month
                )

                posm_num_sum, posm_den_sum = 0, 0
                for posm in posms:
                    if posm.raw_data and "/" in posm.raw_data:
                        try:
                            num, denom = map(int, posm.raw_data.split("/"))
                            posm_num_sum += num
                            posm_den_sum += denom
                        except ValueError:
                            continue

                posm_compliance = (
                    round((posm_num_sum / posm_den_sum) * 100, 2)
                    if posm_den_sum else 0
                )

                total_compliance = (cooler_compliance * (2 / 3)) + (posm_compliance * (1 / 3))
                compliance_volume[month - 1] = round(total_compliance, 2)

# --- Consumer Volumes (Target values) ---
                consumer_volumes = (
                    Restaurant_Target.objects.filter(
                        restaurant__in=gm_restaurants,
                        **totaltargetsfilter,
                        **gmbasefilter,
                        **rsmbasefilter,**restaurantbasefilter,
                        target_type="Consumer",
                        target_status=True
                    )
                    .annotate(month=ExtractMonth("created_at"))   # Extract numeric month (1–12)
                    .values("month")
                    .annotate(total_target=Sum("target_value"))
                    .order_by("month")
                )

                # --- Consumer Actuals (count of checked Consumers) ---
                consumer_targets = (
                    Consumer.objects.filter(
                        restaurant__in=gm_restaurants,
                        **bottlerbasefilter,
                        **gmbasefilter,
                        **rsmbasefilter,**restaurantbasefilter,
                        is_checked=True
                    )
                    .annotate(month=ExtractMonth("created_at"))   # Group by created_at month
                    .values("month")
                    .annotate(total_count=Count("id"))
                    .order_by("month")
                )

                # --- Prepare lists for Chart.js ---
                consumer_month_labels = [calendar.month_abbr[i] for i in range(1, 13)]
                consumer_target = [0] * 12   # Actual consumers (count)
                consumer_volume = [0] * 12   # Planned target (sum of target_value)

                # Fill actuals
                for t in consumer_targets:
                    if t["month"]:
                        consumer_target[t["month"] - 1] = t["total_count"]

                # Fill planned targets
                for v in consumer_volumes:
                    if v["month"]:
                        consumer_volume[v["month"] - 1] = v["total_target"] or 0

        except GM.DoesNotExist:
            try:
                rsm = RSM.objects.get(user_id=user_id)
                rsm_restaurants = rsm.restaurants.all()

                bottlers = Bottler.objects.filter(id=rsm.bottler.id)
                restaurants = Restaurant.objects.filter(id__in=rsm_restaurants.values_list("id", flat=True))
                gms = GM.objects.filter(restaurants__in=rsm_restaurants).distinct()
                rsms = RSM.objects.filter(restaurants__in=rsm_restaurants).distinct()

                bottle_crates_total = Restaurant_Target.objects.filter(
                    restaurant__in=rsm_restaurants,
                    **totaltargetsfilter,
                    **gmbasefilter,
                    **rsmbasefilter,**restaurantbasefilter,
                    target_type="Cases Purchased"
                ).aggregate(total=Sum("target_value"))["total"] or 0

                bottle_crate_target_achieved = Restaurant_Crate_Sales.objects.filter(
                    restaurant__in = rsm_restaurants,
                    **bottlerbasefilter,
                    **gmbasefilter,
                    **rsmbasefilter,**restaurantbasefilter,
                    is_approved = True
                ).aggregate(total=Sum("crate_quantity"))["total"] or 0        

                coolers = Restaurant_Cooler.objects.filter(restaurant__in=rsm_restaurants,is_checked=True,**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)

                cooler_numerator_sum = 0
                cooler_denominator_sum = 0

                for cooler in coolers:
                    if cooler.raw_data and "/" in cooler.raw_data:
                        num, denom = cooler.raw_data.split("/")
                        try:
                            cooler_numerator_sum += int(num)
                            cooler_denominator_sum += int(denom)
                        except ValueError:
                            continue

                cooler_compliance = round((cooler_numerator_sum / cooler_denominator_sum) * 100, 2) if cooler_denominator_sum else 0


                posms = Restaurant_POSM.objects.filter(restaurant__in=rsm_restaurants,is_checked=True,**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter)

                posm_numerator_sum = 0
                posm_denominator_sum = 0

                for posm in posms:
                    if posm.raw_data and "/" in posm.raw_data:
                        num, denom = posm.raw_data.split("/")
                        try:
                            posm_numerator_sum += int(num)
                            posm_denominator_sum += int(denom)
                        except ValueError:
                            continue

                posm_compliance = round((posm_numerator_sum / posm_denominator_sum) * 100, 2) if posm_denominator_sum else 0
 


                total_compliance_target = Restaurant_Target.objects.filter(
                    restaurant__in=rsm_restaurants,
                    **totaltargetsfilter,
                    **gmbasefilter,
                    **rsmbasefilter,**restaurantbasefilter,
                    target_type="Compliance"
                ).aggregate(total=Avg("target_value"))["total"] or 0

                total_compliance_achieved = (cooler_compliance*(2/3))+(posm_compliance*(1/3))

                
                consumer_total = Restaurant_Target.objects.filter(
                    restaurant__in=rsm_restaurants,
                    **totaltargetsfilter,
                    **gmbasefilter,
                    **rsmbasefilter,**restaurantbasefilter,
                    target_type="Consumer"
                ).aggregate(total=Sum("target_value"))["total"] or 0

                consumer_target_achieved = Consumer.objects.filter(
                    restaurant__in=rsm_restaurants,
                    **bottlerbasefilter,
                    **gmbasefilter,
                    **rsmbasefilter,**restaurantbasefilter,
                    is_checked=True
                ).count()


                total_restaurant = rsm_restaurants.count()
                total_bottler = Bottler.objects.filter(id=rsm.bottler.id,**totalbottlerfilter,**bottlergmbase,**bottlerrestaurantbase).count

                if bottle_crates_total > 0:
                    sales_achievement = round((bottle_crate_target_achieved / bottle_crates_total)*100,2)
                if consumer_total > 0:
                    consumer_achievement = round((consumer_target_achieved/consumer_total)*100,2)
                overall_target_achieved = round((cooler_compliance*0.2)+(posm_compliance*0.1)+(sales_achievement*0.5)+(consumer_achievement*0.2),2)
           
                # Sales Graph

                volumes = (
                Restaurant_Target.objects.filter(**totaltargetsfilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter,restaurant__in=rsm_restaurants,target_type="Cases Purchased",target_status=True)
                .annotate(month=ExtractMonth("created_at"))
                .values("month")
                .annotate(total_target=Sum("target_value"))
                .order_by("month")
            )
                
                targets = (
                    Restaurant_Crate_Sales.objects.filter(**bottlerbasefilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter,restaurant__in=rsm_restaurants,is_approved=True)
                    .annotate(month=ExtractMonth("created_at"))
                    .values("month")
                    .annotate(crate_quantity=Sum("crate_quantity"))
                    .order_by("month")
                )


                month_labels = [calendar.month_abbr[i] for i in range(1, 13)]
                sales_target = [0] * 12
                sales_volume = [0] * 12

                for t in targets:
                    sales_target[t["month"] - 1] = round(t["crate_quantity"], 2)

                for v in volumes:
                    sales_volume[v["month"] - 1] = v["total_target"]           
           
# compliance graph data
                compliance_volumes = (
                        Restaurant_Target.objects.filter(**totaltargetsfilter,**gmbasefilter,**rsmbasefilter,**restaurantbasefilter,restaurant__in=rsm_restaurants,target_type="Compliance", target_status=True)
                        .annotate(month=ExtractMonth("created_at"))
                        .values("month")
                        .annotate(total_target=Avg("target_value"))
                        .order_by("month")
                    )

                compliance_month_labels = [calendar.month_abbr[i] for i in range(1, 13)]
                compliance_target = [0] * 12  
                compliance_volume = [0] * 12  

                for v in compliance_volumes:
                    compliance_target[v["month"] - 1] = round(v["total_target"], 2)

                for month in range(1, 13):
                    # --- Coolers ---
                    coolers = Restaurant_Cooler.objects.filter(
                        restaurant__in=rsm_restaurants,
                        **bottlerbasefilter,
                        **gmbasefilter,
                        **rsmbasefilter,**restaurantbasefilter,
                        is_checked=True,
                        created_at__month=month
                    )

                    cooler_num_sum, cooler_den_sum = 0, 0
                    for cooler in coolers:
                        if cooler.raw_data and "/" in cooler.raw_data:
                            try:
                                num, denom = map(int, cooler.raw_data.split("/"))
                                cooler_num_sum += num
                                cooler_den_sum += denom
                            except ValueError:
                                continue

                    cooler_compliance = (
                        round((cooler_num_sum / cooler_den_sum) * 100, 2)
                        if cooler_den_sum else 0
                    )

                    # --- POSMs ---
                    posms = Restaurant_POSM.objects.filter(
                        restaurant__in=rsm_restaurants,
                        **bottlerbasefilter,
                        **gmbasefilter,
                        **rsmbasefilter,**restaurantbasefilter,
                        is_checked=True,
                        created_at__month=month
                    )

                    posm_num_sum, posm_den_sum = 0, 0
                    for posm in posms:
                        if posm.raw_data and "/" in posm.raw_data:
                            try:
                                num, denom = map(int, posm.raw_data.split("/"))
                                posm_num_sum += num
                                posm_den_sum += denom
                            except ValueError:
                                continue

                    posm_compliance = (
                        round((posm_num_sum / posm_den_sum) * 100, 2)
                        if posm_den_sum else 0
                    )

                    total_compliance = (cooler_compliance * (2 / 3)) + (posm_compliance * (1 / 3))
                    compliance_volume[month - 1] = round(total_compliance, 2)




# --- Consumer Volumes (Target values) ---
                    consumer_volumes = (
                        Restaurant_Target.objects.filter(
                            restaurant__in=rsm_restaurants,
                            **totaltargetsfilter,
                            **gmbasefilter,
                            **rsmbasefilter,**restaurantbasefilter,
                            target_type="Consumer",
                            target_status=True
                        )
                        .annotate(month=ExtractMonth("created_at"))   # Extract numeric month (1–12)
                        .values("month")
                        .annotate(total_target=Sum("target_value"))
                        .order_by("month")
                    )

                    # --- Consumer Actuals (count of checked Consumers) ---
                    consumer_targets = (
                        Consumer.objects.filter(
                            **bottlerbasefilter,
                            **gmbasefilter,
                            **rsmbasefilter,**restaurantbasefilter,
                            restaurant__in=rsm_restaurants,
                            is_checked=True
                        )
                        .annotate(month=ExtractMonth("created_at"))   # Group by created_at month
                        .values("month")
                        .annotate(total_count=Count("id"))
                        .order_by("month")
                    )

                    # --- Prepare lists for Chart.js ---
                    consumer_month_labels = [calendar.month_abbr[i] for i in range(1, 13)]
                    consumer_target = [0] * 12   # Actual consumers (count)
                    consumer_volume = [0] * 12   # Planned target (sum of target_value)

                    # Fill actuals
                    for t in consumer_targets:
                        if t["month"]:
                            consumer_target[t["month"] - 1] = t["total_count"]

                    # Fill planned targets
                    for v in consumer_volumes:
                        if v["month"]:
                            consumer_volume[v["month"] - 1] = v["total_target"] or 0

            except RSM.DoesNotExist:
                bottle_crates_total = 0





        
    total_compliance_target = round(total_compliance_target,2)

    total_compliance_achieved = round(total_compliance_achieved)

    return render(
        request,
        "dashboard.html",
        {"user":user,"username": username,"bottlers":bottlers,"restaurants":restaurants,"gms":gms,"rsms":rsms,"bottle_crates_total": bottle_crates_total ,"bottle_crate_target_achieved":bottle_crate_target_achieved,"total_compliance_target":total_compliance_target,"total_compliance_achieved":total_compliance_achieved,"total_bottler":total_bottler,"total_restaurant":total_restaurant,"consumer_target_achieved":consumer_target_achieved,"consumer_total":consumer_total,"overall_target_achieved":overall_target_achieved, "labels": json.dumps(month_labels),
    "sales_target": json.dumps(sales_target),
    "sales_volume": json.dumps(sales_volume),"compliance_month_labels":json.dumps(compliance_month_labels),"compliance_target": json.dumps(compliance_target),  # Target % values
    "compliance_volume": json.dumps(compliance_volume),"consumer_month_labels":json.dumps(consumer_month_labels),"consumer_target":consumer_target,"consumer_volume":consumer_volume}
    )


def sales_data_date(request):
    user_id = request.session.get("user_id")
    user = User.objects.get(id=user_id)
    month_name = request.GET.get("month")
    filters = {}


    if not month_name:
        return JsonResponse({"error": "Month is required"}, status=400)

    try:
        month_num = list(calendar.month_abbr).index(month_name)
    except ValueError:
        return JsonResponse({"error": "Invalid month"}, status=400)
    import datetime

    year = datetime.date.today().year
    days_in_month = calendar.monthrange(year, month_num)[1]



    if user.is_superuser:
        daily_targets = (
            Restaurant_Target.objects.filter(
                target_type="Cases Purchased", target_status=True,
                created_at__year=year, created_at__month=month_num, 
            )
            .annotate(day=ExtractDay("created_at"))
            .values("day")
            .annotate(total_target=Sum("target_value"))
            .order_by("day")
        )

        daily_sales = (
            Restaurant_Crate_Sales.objects.filter(
                is_approved=True, created_at__year=year,
                created_at__month=month_num, 
            )
            .annotate(day=ExtractDay("created_at"))
            .values("day")
            .annotate(crate_quantity=Sum("crate_quantity"))
            .order_by("day")
        )

        day_labels = [str(i) for i in range(1, days_in_month + 1)]
        sales_target = [0] * days_in_month
        sales_volume = [0] * days_in_month

        for t in daily_sales:
            sales_target[t["day"] - 1] = round(t["crate_quantity"], 2)
        for v in daily_targets:
            sales_volume[v["day"] - 1] = v["total_target"]

    else:
        try:
            gm = GM.objects.get(user_id=user_id)
            gm_restaurants = gm.restaurants.all()

            daily_targets = (
            Restaurant_Target.objects.filter(
                restaurant__in=gm_restaurants,
                target_type="Cases Purchased", target_status=True,
                created_at__year=year, created_at__month=month_num, 
            )
            .annotate(day=ExtractDay("created_at"))
            .values("day")
            .annotate(total_target=Sum("target_value"))
            .order_by("day")
        )

            daily_sales = (
                Restaurant_Crate_Sales.objects.filter(
                    restaurant__in=gm_restaurants,
                    is_approved=True, created_at__year=year,
                    created_at__month=month_num, 
                )
                .annotate(day=ExtractDay("created_at"))
                .values("day")
                .annotate(crate_quantity=Sum("crate_quantity"))
                .order_by("day")
            )

            day_labels = [str(i) for i in range(1, days_in_month + 1)]
            sales_target = [0] * days_in_month
            sales_volume = [0] * days_in_month

            for t in daily_sales:
                sales_target[t["day"] - 1] = round(t["crate_quantity"], 2)
            for v in daily_targets:
                sales_volume[v["day"] - 1] = v["total_target"]

        except GM.DoesNotExist:
            try:
                rsm = RSM.objects.get(user_id=user_id)
                rsm_restaurants = rsm.restaurants.all()

                daily_targets = (
                Restaurant_Target.objects.filter(
                    restaurant__in=rsm_restaurants,
                    target_type="Cases Purchased", target_status=True,
                    created_at__year=year, created_at__month=month_num, 
                )
                .annotate(day=ExtractDay("created_at"))
                .values("day")
                .annotate(total_target=Sum("target_value"))
                .order_by("day")
            )

                daily_sales = (
                    Restaurant_Crate_Sales.objects.filter(
                        restaurant__in=rsm_restaurants,
                        is_approved=True, created_at__year=year,
                        created_at__month=month_num, 
                    )
                    .annotate(day=ExtractDay("created_at"))
                    .values("day")
                    .annotate(crate_quantity=Sum("crate_quantity"))
                    .order_by("day")
                )

                day_labels = [str(i) for i in range(1, days_in_month + 1)]
                sales_target = [0] * days_in_month
                sales_volume = [0] * days_in_month

                for t in daily_sales:
                    sales_target[t["day"] - 1] = round(t["crate_quantity"], 2)
                for v in daily_targets:
                    sales_volume[v["day"] - 1] = v["total_target"]  
            except RSM.DoesNotExist:
                day_labels = 0
                sales_target = 0
                sales_volume = 0               




    return JsonResponse({
        "labels": day_labels,
        "sales_target": sales_target,
        "sales_volume": sales_volume,
    })


def compliance_data_date(request):
    user_id = request.session.get("user_id")
    user = User.objects.get(id=user_id)

    month_name = request.GET.get("month")
    if not month_name:
        return JsonResponse({"error": "Month is required"}, status=400)

    try:
        month_num = list(calendar.month_abbr).index(month_name)
    except ValueError:
        return JsonResponse({"error": "Invalid month"}, status=400)

    import datetime
    year = datetime.date.today().year
    days_in_month = calendar.monthrange(year, month_num)[1]

    # --- Labels for chart ---
    day_labels = [str(i) for i in range(1, days_in_month + 1)]
    compliance_target = [0] * days_in_month
    compliance_volume = [0] * days_in_month

    # --- Select restaurants depending on role ---
    if user.is_superuser:
        restaurants = Restaurant.objects.all()
    else:
        try:
            gm = GM.objects.get(user_id=user_id)
            restaurants = gm.restaurants.all()
        except GM.DoesNotExist:
            try:
                rsm = RSM.objects.get(user_id=user_id)
                restaurants = rsm.restaurants.all()
            except RSM.DoesNotExist:
                return JsonResponse({
                    "labels": [],
                    "compliance_target": [],
                    "compliance_volume": []
                })

    # --- Compliance Target (Planned) per Day ---
    compliance_targets = (
        Restaurant_Target.objects.filter(
            restaurant__in=restaurants,
            target_type="Compliance",
            target_status=True,
            created_at__year=year,
            created_at__month=month_num,
        )
        .annotate(day=ExtractDay("created_at"))
        .values("day")
        .annotate(avg_target=Avg("target_value"))
        .order_by("day")
    )

    for t in compliance_targets:
        compliance_target[t["day"] - 1] = round(t["avg_target"], 2) if t["avg_target"] else 0

    # --- Compliance Actual (Achieved) per Day ---
    for day in range(1, days_in_month + 1):
        # Coolers
        coolers = Restaurant_Cooler.objects.filter(
            restaurant__in=restaurants,
            is_checked=True,
            created_at__year=year,
            created_at__month=month_num,
            created_at__day=day,
        )
        cooler_num_sum, cooler_den_sum = 0, 0
        for cooler in coolers:
            if cooler.raw_data and "/" in cooler.raw_data:
                try:
                    num, denom = map(int, cooler.raw_data.split("/"))
                    cooler_num_sum += num
                    cooler_den_sum += denom
                except ValueError:
                    continue
        cooler_compliance = (
            round((cooler_num_sum / cooler_den_sum) * 100, 2)
            if cooler_den_sum else 0
        )

        # POSMs
        posms = Restaurant_POSM.objects.filter(
            restaurant__in=restaurants,
            is_checked=True,
            created_at__year=year,
            created_at__month=month_num,
            created_at__day=day,
        )
        posm_num_sum, posm_den_sum = 0, 0
        for posm in posms:
            if posm.raw_data and "/" in posm.raw_data:
                try:
                    num, denom = map(int, posm.raw_data.split("/"))
                    posm_num_sum += num
                    posm_den_sum += denom
                except ValueError:
                    continue
        posm_compliance = (
            round((posm_num_sum / posm_den_sum) * 100, 2)
            if posm_den_sum else 0
        )

        # Weighted total compliance
        total_compliance = (cooler_compliance * (2 / 3)) + (posm_compliance * (1 / 3))
        compliance_volume[day - 1] = round(total_compliance, 2)

    return JsonResponse({
        "labels": day_labels,
        "compliance_target": compliance_target,
        "compliance_volume": compliance_volume,
    })



def consumer_data_date(request):

    user_id = request.session.get("user_id")
    user = User.objects.get(id=user_id)

    month_name = request.GET.get("month")

    if not month_name:
        return JsonResponse({"error": "Month is required"}, status=400)

    try:
        month_num = list(calendar.month_abbr).index(month_name)
    except ValueError:
        return JsonResponse({"error": "Invalid month"}, status=400)
    import datetime
    year = datetime.date.today().year
    days_in_month = calendar.monthrange(year, month_num)[1]


    if user.is_superuser:

        consumer_volumes = (
                Restaurant_Target.objects.filter(
                    target_type="Consumer",
                    target_status=True,
                    created_at__year=year,
                    created_at__month=month_num,
                )
                .annotate(day=ExtractDay("created_at"))
                .values("day")
                .annotate(total_target=Sum("target_value"))
                .order_by("day")
            )

        consumer_targets = (
            Consumer.objects.filter(
                is_checked=True,
                created_at__year=year,
                created_at__month=month_num,
            )
            .annotate(day=ExtractDay("created_at"))
            .values("day")
            .annotate(total_count=Count("id"))
            .order_by("day")
        )

        # Prepare lists
        day_labels = [str(i) for i in range(1, days_in_month + 1)]
        consumer_target = [0] * days_in_month
        consumer_volume = [0] * days_in_month

        for t in consumer_targets:
            if t["day"]:
                consumer_target[t["day"] - 1] = t["total_count"]

        for v in consumer_volumes:
            if v["day"]:
                consumer_volume[v["day"] - 1] = v["total_target"] or 0

    else:
        try:
            gm = GM.objects.get(user_id=user_id)
            gm_restaurants = gm.restaurants.all()     
            consumer_volumes = (
                Restaurant_Target.objects.filter(
                    restaurant__in=gm_restaurants,
                    target_type="Consumer",
                    target_status=True,
                    created_at__year=year,
                    created_at__month=month_num,
                )
                .annotate(day=ExtractDay("created_at"))
                .values("day")
                .annotate(total_target=Sum("target_value"))
                .order_by("day")
            )

            # Actual (count of checked Consumers)
            consumer_targets = (
                Consumer.objects.filter(
                    restaurant__in=gm_restaurants,
                    is_checked=True,
                    created_at__year=year,
                    created_at__month=month_num,
                )
                .annotate(day=ExtractDay("created_at"))
                .values("day")
                .annotate(total_count=Count("id"))
                .order_by("day")
            )

            # Prepare lists
            day_labels = [str(i) for i in range(1, days_in_month + 1)]
            consumer_target = [0] * days_in_month
            consumer_volume = [0] * days_in_month

            for t in consumer_targets:
                if t["day"]:
                    consumer_target[t["day"] - 1] = t["total_count"]

            for v in consumer_volumes:
                if v["day"]:
                    consumer_volume[v["day"] - 1] = v["total_target"] or 0

        except GM.DoesNotExist:
            try:
                rsm = RSM.objects.get(user_id=user_id)
                rsm_restaurants = rsm.restaurants.all()     


                consumer_volumes = (
                    Restaurant_Target.objects.filter(
                        restaurant__in=rsm_restaurants,
                        target_type="Consumer",
                        target_status=True,
                        created_at__year=year,
                        created_at__month=month_num,
                    )
                    .annotate(day=ExtractDay("created_at"))
                    .values("day")
                    .annotate(total_target=Sum("target_value"))
                    .order_by("day")
                )

                consumer_targets = (
                    Consumer.objects.filter(
                        restaurant__in=rsm_restaurants,
                        is_checked=True,
                        created_at__year=year,
                        created_at__month=month_num,
                    )
                    .annotate(day=ExtractDay("created_at"))
                    .values("day")
                    .annotate(total_count=Count("id"))
                    .order_by("day")
                )


                # Prepare lists
                day_labels = [str(i) for i in range(1, days_in_month + 1)]
                consumer_target = [0] * days_in_month
                consumer_volume = [0] * days_in_month

                for t in consumer_targets:
                    if t["day"]:
                        consumer_target[t["day"] - 1] = t["total_count"]

                for v in consumer_volumes:
                    if v["day"]:
                        consumer_volume[v["day"] - 1] = v["total_target"] or 0

            except RSM.DoesNotExist:
                day_labels = 0
                consumer_target = 0
                consumer_volume = 0  

    print(consumer_target)
    print(consumer_volume)


    return JsonResponse({
        "labels": day_labels,
        "consumer_target": consumer_target,
        "consumer_volume": consumer_volume,
    })




def get_gms_by_bottler(request, bottler_id):
    gms = GM.objects.filter(bottler_id=bottler_id).values("id", "gm_name")
    return JsonResponse(list(gms), safe=False)

def get_rsms_by_bottler(request, bottler_id):
    rsms = RSM.objects.filter(bottler_id=bottler_id).values("id", "rsm_name")
    return JsonResponse(list(rsms), safe=False)

def get_restaurants_by_bottler(request, bottler_id):
    restaurants = Restaurant.objects.filter(bottler=bottler_id).values("id", "restaurant_name")
    return JsonResponse(list(restaurants), safe=False)

def get_rsms_by_gm(request, gm_id):
    rsms = RSM.objects.filter(tagged_gms=gm_id).values("id", "rsm_name")
    return JsonResponse(list(rsms), safe=False)

def get_restaurants_by_gm(request, gm_id):
    restaurants = Restaurant.objects.filter(gms__id=gm_id).values("id", "restaurant_name")
    return JsonResponse(list(restaurants), safe=False)

def get_restaurants_by_rsm(request, rsm_id):
    restaurants = Restaurant.objects.filter(rsms__id=rsm_id).values("id", "restaurant_name")
    return JsonResponse(list(restaurants), safe=False)



@login_required(login_url='/dashboard/login/')

def bottler(request):
    user_id = request.session.get('user_id')
    user = User.objects.get(id=user_id)
    username = user.username
    gms = []
    rsms = []

    search_query = request.GET.get("q", "").strip()  # ✅ Search keyword
    import datetime
    today = datetime.date.today()
    current_month = today.month
    current_year = today.year
    if user.is_superuser:
        gms = GM.objects.all()
        rsms = RSM.objects.all()

        # ✅ Apply search filter
        if search_query:
            gms = gms.filter(
                Q(gm_name__icontains=search_query)
                | Q(gm_phone_number__icontains=search_query)
                | Q(bottler__bottler_name__icontains=search_query)
            )
            rsms = rsms.filter(
                Q(rsm_name__icontains=search_query)
                | Q(rsm_phone_number__icontains=search_query)
                | Q(bottler__bottler_name__icontains=search_query)
            )

        for gm in gms:
            gm_bottle_own_total = Restaurant_Target.objects.filter(
                restaurant__in=gm.restaurants.all(),
                target_type="Cases Purchased",
                created_at__month=current_month,
                created_at__year=current_year
            ).aggregate(total=Sum("target_value"))["total"] or 0

            gm_consumer_own_total = Restaurant_Target.objects.filter(
                restaurant__in=gm.restaurants.all(),
                target_type="Consumer",
                created_at__month=current_month,
                created_at__year=current_year
            ).aggregate(total=Sum("target_value"))["total"] or 0

            gm_compliance_own_total = Restaurant_Target.objects.filter(
                restaurant__in=gm.restaurants.all(),
                target_type="Compliance",
                created_at__month=current_month,
                created_at__year=current_year
            ).aggregate(total=Avg("target_value"))["total"] or 0


            gm_bottle_from_rsms = 0
            gm_consumer_from_rsms = 0
            gm_compliance_from_rsms = 0

            tagged_rsms = RSM.objects.filter(tagged_gms=gm)

            for rsm in tagged_rsms:
                rsm_bottle_total = Restaurant_Target.objects.filter(
                    restaurant__in=rsm.restaurants.all(),
                    target_type="Cases Purchased",
                created_at__month=current_month,
                created_at__year=current_year
                ).aggregate(total=Sum("target_value"))["total"] or 0
                rsm.rsm_bottle_crate_target = rsm_bottle_total
                gm_bottle_from_rsms += rsm_bottle_total

                rsm_consumer_total = Restaurant_Target.objects.filter(
                    restaurant__in=rsm.restaurants.all(),
                    target_type="Consumer",
                created_at__month=current_month,
                created_at__year=current_year                    
                ).aggregate(total=Sum("target_value"))["total"] or 0
                rsm.rsm_consumer_target = rsm_consumer_total
                gm_consumer_from_rsms += rsm_consumer_total

                rsm_compliance_total = Restaurant_Target.objects.filter(
                    restaurant__in=rsm.restaurants.all(),
                    target_type="Compliance",
                created_at__month=current_month,
                created_at__year=current_year
                ).aggregate(total=Avg("target_value"))["total"] or 0
                rsm.rsm_compliance_target = rsm_compliance_total
                gm_compliance_from_rsms += rsm_compliance_total

            gm.gm_bottle_crate_target = gm_bottle_own_total
            gm.gm_consumer_target = gm_consumer_own_total
            gm.gm_compliance_target = round(gm_compliance_own_total,2)

        for rsm in rsms:
            rsm_bottle_total = Restaurant_Target.objects.filter(
                restaurant__in=rsm.restaurants.all(),
                target_type="Cases Purchased",
                created_at__month=current_month,
                created_at__year=current_year
            ).aggregate(total=Sum("target_value"))["total"] or 0
            rsm.rsm_bottle_crate_target = rsm_bottle_total

            rsm_consumer_total = Restaurant_Target.objects.filter(
                restaurant__in=rsm.restaurants.all(),
                target_type="Consumer",
                created_at__month=current_month,
                created_at__year=current_year
            ).aggregate(total=Sum("target_value"))["total"] or 0
            rsm.rsm_consumer_target = rsm_consumer_total

            rsm.rsm_compliance_total = Restaurant_Target.objects.filter(
                restaurant__in=rsm.restaurants.all(),
                target_type="Compliance"
            ).aggregate(total=Avg("target_value"))["total"] or 0
            rsm.rsm_compliance_target = round(rsm.rsm_compliance_total,2)

    else:
        try:
            gm = GM.objects.get(user_id=user_id)
            gms = [gm]
            rsms = RSM.objects.filter(tagged_gms=gm)

            # ✅ Apply search filter
            if search_query:
                gms = [gm for gm in gms if search_query.lower() in gm.gm_name.lower() or search_query.lower() in gm.gm_phone_number.lower() or search_query.lower() in gm.bottler.bottler_name.lower()]
                rsms = rsms.filter(
                    Q(rsm_name__icontains=search_query)
                    | Q(rsm_phone_number__icontains=search_query)
                    | Q(bottler__bottler_name__icontains=search_query)
                )

            gm_bottle_own_total = Restaurant_Target.objects.filter(
                restaurant__in=gm.restaurants.all(),
                target_type="Cases Purchased",
                created_at__month=current_month,
                created_at__year=current_year
            ).aggregate(total=Sum("target_value"))["total"] or 0

            gm_consumer_own_total = Restaurant_Target.objects.filter(
                restaurant__in=gm.restaurants.all(),
                target_type="Consumer",
                created_at__month=current_month,
                created_at__year=current_year
            ).aggregate(total=Sum("target_value"))["total"] or 0

            gm_compliance_own_total = Restaurant_Target.objects.filter(
                restaurant__in=gm.restaurants.all(),
                target_type="Compliance",
                created_at__month=current_month,
                created_at__year=current_year
            ).aggregate(total=Avg("target_value"))["total"] or 0


            gm_bottle_from_rsms = 0
            gm_consumer_from_rsms = 0
            gm_compliance_own_total = 0

            for rsm in rsms:
                rsm_bottle_total = Restaurant_Target.objects.filter(
                    restaurant__in=rsm.restaurants.all(),
                    target_type="Cases Purchased",
                created_at__month=current_month,
                created_at__year=current_year
                ).aggregate(total=Sum("target_value"))["total"] or 0
                rsm.rsm_bottle_crate_target = rsm_bottle_total
                gm_bottle_from_rsms += rsm_bottle_total

                rsm_consumer_total = Restaurant_Target.objects.filter(
                    restaurant__in=rsm.restaurants.all(),
                    target_type="Consumer",
                created_at__month=current_month,
                created_at__year=current_year
                ).aggregate(total=Sum("target_value"))["total"] or 0
                rsm.rsm_consumer_target = rsm_consumer_total
                gm_consumer_from_rsms += rsm_consumer_total

                rsm_compliance_total = Restaurant_Target.objects.filter(
                    restaurant__in=rsm.restaurants.all(),
                    target_type="Compliance",
                created_at__month=current_month,
                created_at__year=current_year
                ).aggregate(total=Avg("target_value"))["total"] or 0
                rsm.rsm_compliance_target = rsm_compliance_total
                print(rsm_compliance_total)
                gm_compliance_own_total += rsm_compliance_total



            gm.gm_bottle_crate_target = gm_bottle_own_total 
            gm.gm_consumer_target = gm_consumer_own_total 
            gm.gm_compliance_target = round(gm_compliance_own_total,2)

        except GM.DoesNotExist:
            gms = []
            rsms = []

        try:
            rsm = RSM.objects.get(user_id=user_id)
            rsms = [rsm]

            # ✅ Apply search filter
            if search_query:
                rsms = [rsm for rsm in rsms if search_query.lower() in rsm.rsm_name.lower() or search_query.lower() in rsm.rsm_phone_number.lower() or search_query.lower() in rsm.bottler.bottler_name.lower()]

            rsm_bottle_total = Restaurant_Target.objects.filter(
                restaurant__in=rsm.restaurants.all(),
                target_type="Cases Purchased",
                created_at__month=current_month,
                created_at__year=current_year
            ).aggregate(total=Sum("target_value"))["total"] or 0
            rsm.rsm_bottle_crate_target = rsm_bottle_total

            rsm_consumer_total = Restaurant_Target.objects.filter(
                restaurant__in=rsm.restaurants.all(),
                target_type="Consumer",
                created_at__month=current_month,
                created_at__year=current_year
            ).aggregate(total=Sum("target_value"))["total"] or 0
            rsm.rsm_consumer_target = rsm_consumer_total

            
            rsm_compliance_total = Restaurant_Target.objects.filter(
                restaurant__in=rsm.restaurants.all(),
                target_type="Compliance",
                created_at__month=current_month,
                created_at__year=current_year
            ).aggregate(total=Avg("target_value"))["total"] or 0
            rsm.rsm_compliance_target = round(rsm_compliance_total,2)

        except RSM.DoesNotExist:
            pass

    # ✅ Combine GM and RSM into one list for pagination
    all_records = list(gms) + list(rsms)

    # ✅ Apply pagination
    paginator = Paginator(all_records, 10)  # show 10 per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "bottlers.html",
        {"user": user, "username": username, "page_obj": page_obj, "search_query": search_query},
    )


@login_required(login_url='/dashboard/login/')
def details_gm(request, gm_id):
    user_id = request.session.get('user_id')
    user = User.objects.get(id=user_id)
    username = user.username    

    gm = GM.objects.get(id=gm_id)

    if user.is_superuser:
        all_bottlers = Bottler.objects.all()
    else:
        all_bottlers = None

    bottle_crates_total = Restaurant_Target.objects.filter(
        restaurant__in=gm.restaurants.all(),
        target_type="Cases Purchased"
    ).aggregate(total=Sum("target_value"))["total"] or 0

    consumer_total = Restaurant_Target.objects.filter(
        restaurant__in=gm.restaurants.all(),
        target_type="Consumer"
    ).aggregate(total=Sum("target_value"))["total"] or 0    

    compliance_total = Restaurant_Target.objects.filter(
        restaurant__in=gm.restaurants.all(),
        target_type="Compliance"
    ).aggregate(total=Avg("target_value"))["total"] or 0    


    return render(
        request,
        "details-gm.html",
        {
            "gm": gm,
            "all_bottlers": all_bottlers,
            "username": username,
            "user":user,
            "bottle_crates_total": bottle_crates_total,
            "consumer_total": consumer_total,
            "compliance_total": round(compliance_total,2)
        }
    )

@login_required(login_url='/dashboard/login/')
def details_rsm(request,rsm_id):
    user_id = request.session.get('user_id')
    user = User.objects.get(id=user_id)
    username = user.username    

    rsm = RSM.objects.get(id=rsm_id)

    if user.is_superuser:
        all_bottlers = Bottler.objects.all()
    else:
        all_bottlers = None

    bottle_crates_total = Restaurant_Target.objects.filter(
        restaurant__in=rsm.restaurants.all(),
        target_type="Cases Purchased"
    ).aggregate(total=Sum("target_value"))["total"] or 0

    consumer_total = Restaurant_Target.objects.filter(
        restaurant__in=rsm.restaurants.all(),
        target_type="Consumer"
    ).aggregate(total=Sum("target_value"))["total"] or 0

    compliance_total = Restaurant_Target.objects.filter(
        restaurant__in=rsm.restaurants.all(),
        target_type="Compliance"
    ).aggregate(total=Avg("target_value"))["total"] or 0    

    return render(request,"details-rsm.html",
                  
                     {"rsm": rsm,
            "all_bottlers": all_bottlers,
            "username": username,
            "user":user,
            "bottle_crates_total": bottle_crates_total,
            "consumer_total": consumer_total, 
            "compliance_total":round(compliance_total,2)        
                              }        )

@login_required(login_url='/dashboard/login/')
@csrf_exempt
def edit_rsm_api(request, rsm_id):
    if request.method != "PUT":
        return JsonResponse({"success": False, "message": "Only PUT allowed"}, status=405)

    # --- JWT Auth ---
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid token: {e}"}, status=401)

    # --- Parse body ---
    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON"}, status=400)

    rsmname = data.get("name", "").strip()
    rsmcontactNo = data.get("contactNo")
    rsmbottler = data.get("bottler")
    crate_target = data.get("crate_target")
    consumer_target = data.get("consumer_target")
    compliance_target = data.get("compliance_target")

    if not rsmname or rsmcontactNo is None:
        return JsonResponse({"success": False, "message": "Name and Contact no are required"}, status=400)

    if not re.fullmatch(r"03\d{9}", rsmcontactNo or ""):
        return JsonResponse({
            "success": False,
            "message": "Contact number must start with 03 and be 11 digits long"
        }, status=400)
    # --- Fetch RSM instance ---
    try:
        rsm = RSM.objects.get(id=rsm_id)
    except RSM.DoesNotExist:
        return JsonResponse({"success": False, "message": "RSM not found"}, status=404)

    # --- Update always-allowed fields ---
    rsm.rsm_name = rsmname
    rsm.rsm_phone_number = rsmcontactNo

    # --- Update restricted fields (only staff/superuser) ---
    if user.is_superuser or user.is_staff:
        if rsmbottler:
            try:
                bottler = Bottler.objects.get(id=rsmbottler)
                rsm.bottler = bottler
            except Bottler.DoesNotExist:
                return JsonResponse({"success": False, "message": "Invalid bottler"}, status=400)

        # Update or create targets (example logic — adjust as needed)
        if crate_target is not None:
            Restaurant_Target.objects.update_or_create(
                restaurant=rsm.restaurants.first(),
                target_type="Cases Purchased",
                defaults={"target_value": crate_target, "target_status": True}
            )
        if consumer_target is not None:
            Restaurant_Target.objects.update_or_create(
                restaurant=rsm.restaurants.first(),
                target_type="Consumer",
                defaults={"target_value": consumer_target, "target_status": True}
            )
        if compliance_target is not None:
            Restaurant_Target.objects.update_or_create(
                restaurant=rsm.restaurants.first(),
                target_type="Compliance",
                defaults={"target_value": compliance_target, "target_status": True}
            )
    else:
        # If non-staff tries to update restricted fields → reject
        if rsmbottler or crate_target or consumer_target or compliance_target:
            return JsonResponse({
                "success": False,
                "message": "You don’t have permission to edit bottler or targets"
            }, status=403)

    # --- Save RSM ---
    rsm.save()

    return JsonResponse({
        "success": True,
        "message": "RSM updated successfully",
        "id": rsm.id,
    }, status=200)


@login_required(login_url='/dashboard/login/')
@csrf_exempt
def edit_gm_api(request, gm_id):
    if request.method != "PUT":
        return JsonResponse({"success": False, "message": "Only PUT allowed"}, status=405)

    # --- JWT Auth ---
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid token: {e}"}, status=401)

    # --- Parse body ---
    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON"}, status=400)

    gmname = data.get("name", "").strip()
    gmcontactNo = data.get("contactNo")
    gmbottler = data.get("bottler")
    crate_target = data.get("crate_target")
    consumer_target = data.get("consumer_target")
    compliance_target = data.get("compliance_target")

    if not gmname or gmcontactNo is None:
        return JsonResponse({"success": False, "message": "Name and Contact no are required"}, status=400)

    if not re.fullmatch(r"03\d{9}", gmcontactNo or ""):
        return JsonResponse({
            "success": False,
            "message": "Contact number must start with 03 and be 11 digits long"
        }, status=400)



    # --- Fetch RSM instance ---
    try:
        gm = GM.objects.get(id=gm_id)
    except RSM.DoesNotExist:
        return JsonResponse({"success": False, "message": "GM not found"}, status=404)

    # --- Update always-allowed fields ---
    gm.gm_name = gmname
    gm.gm_phone_number = gmcontactNo

    # --- Update restricted fields (only staff/superuser) ---
    if user.is_superuser or user.is_staff:
        if gmbottler:
            try:
                bottler = Bottler.objects.get(id=gmbottler)
                gm.bottler = bottler
            except Bottler.DoesNotExist:
                return JsonResponse({"success": False, "message": "Invalid bottler"}, status=400)

        # Update or create targets (example logic — adjust as needed)
        if crate_target is not None:
            Restaurant_Target.objects.update_or_create(
                restaurant=gm.restaurants.first(),
                target_type="Cases Purchased",
                defaults={"target_value": crate_target, "target_status": True}
            )
        if consumer_target is not None:
            Restaurant_Target.objects.update_or_create(
                restaurant=gm.restaurants.first(),
                target_type="Consumer",
                defaults={"target_value": consumer_target, "target_status": True}
            )
        if compliance_target is not None:
            Restaurant_Target.objects.update_or_create(
                restaurant=gm.restaurants.first(),
                target_type="Compliance",
                defaults={"target_value": compliance_target, "target_status": True}
            )
    else:
        # If non-staff tries to update restricted fields → reject
        if gmbottler or crate_target or consumer_target or compliance_target:
            return JsonResponse({
                "success": False,
                "message": "You don’t have permission to edit bottler or targets"
            }, status=403)

    # --- Save RSM ---
    gm.save()

    return JsonResponse({
        "success": True,
        "message": "GM updated successfully",
        "id": gm.id,
    }, status=200)


def delete_gm_api(request, gm_id):
    if request.method != "DELETE":
        return JsonResponse({"success": False, "message": "Only DELETE allowed"}, status=405)

    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid token: {e}"}, status=401)

    if not (user.is_superuser or user.is_staff):
        return JsonResponse({"success": False, "message": "You do not have permission to delete GM"}, status=403)

    try:
        gm = GM.objects.get(id=gm_id)
        linked_user = gm.user 
        gm.delete()

        if linked_user:
            linked_user.delete()
        return JsonResponse({"success": True, "message": "GM deleted successfully"})
    except GM.DoesNotExist:
        return JsonResponse({"success": False, "message": "GM not found"}, status=404)

def delete_rsm_api(request, rsm_id):
    if request.method != "DELETE":
        return JsonResponse({"success": False, "message": "Only DELETE allowed"}, status=405)

    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid token: {e}"}, status=401)

    if not (user.is_superuser or user.is_staff):
        return JsonResponse({"success": False, "message": "You do not have permission to delete RSM"}, status=403)

    try:
        rsm = RSM.objects.get(id=rsm_id)
        linked_user = rsm.user  
        rsm.delete() 

        if linked_user:
            linked_user.delete()  
        return JsonResponse({"success": True, "message": "RSM deleted successfully"})
    except RSM.DoesNotExist:
        return JsonResponse({"success": False, "message": "RSM not found"}, status=404)





@login_required(login_url='/dashboard/login/')
def restaurant(request):
    user_id = request.session.get("user_id")
    user = User.objects.get(id=user_id)
    username = user.username

    if user.is_superuser:
        restaurants_qs = Restaurant.objects.all()
        manager = Manager.objects.all()
    else:
        try:
            gm = GM.objects.get(user_id=user_id)
            restaurants_qs = gm.restaurants.all()
        except GM.DoesNotExist:
            try:
                rsm = RSM.objects.get(user_id=user_id)
                restaurants_qs = rsm.restaurants.all()
            except RSM.DoesNotExist:
                restaurants_qs = Restaurant.objects.none()

    # ✅ Search functionality
    query = request.GET.get("q")  # from ?q=searchterm
    if query:
        restaurants_qs = restaurants_qs.filter(
            Q(restaurant_name__icontains=query) | Q(restaurant_category__icontains=query) 
        )

    for res in restaurants_qs:
        res.gm_list = res.gms.all()
        res.rsm_list = res.rsms.all()

    page_number = request.GET.get("page", 1)
    paginator = Paginator(restaurants_qs, 10)
    page_obj = paginator.get_page(page_number)


    restaurant_coolers = Restaurant_Cooler.objects.filter(
    is_checked=True,
    is_manually=True
).order_by("-created_at")

    cooler_list = []

    for cooler in restaurant_coolers:
        compliance = 0
        if cooler.raw_data and "/" in cooler.raw_data:
            try:
                num, denom = cooler.raw_data.split("/")
                num = int(num)
                denom = int(denom)
                compliance = round((num / denom) * 100, 2) if denom else 0
            except ValueError:
                compliance = 0

        # determine status text
        if cooler.is_manually:
            status = "Requested"
        elif hasattr(cooler, "is_changed") and cooler.is_changed:  # fallback if field exists
            status = "Changed"
        else:
            status = "-"

        cooler_list.append({
            "id": cooler.id,
            "cooler_type": cooler.cooler_type,
            "raw_data": cooler.raw_data,
            "compliance": compliance,
            "status": status,
            "image": cooler.image,
            "created_at": cooler.created_at.strftime("%Y-%m-%d %H:%M"),  # cleaner format
        })



    restaurant_posm = Restaurant_POSM.objects.filter(
        is_checked=True,
        is_manually=True
    ).order_by("-created_at")

    posm_list = []

    for posm in restaurant_posm:
        compliance = 0
        if posm.raw_data and "/" in posm.raw_data:
            try:
                num, denom = posm.raw_data.split("/")
                num = int(num)
                denom = int(denom)
                compliance = round((num / denom) * 100, 2) if denom else 0
            except ValueError:
                compliance = 0

        # determine status text
        if posm.is_manually:
            status = "Requested"
        elif hasattr(posm, "is_changed") and posm.is_changed:  # fallback if field exists
            status = "Changed"
        else:
            status = "-"

        posm_list.append({
            "id": posm.id,
            "posm_type": posm.posm_type,
            "raw_data": posm.raw_data,
            "compliance": compliance,
            "status": status,
            "image": posm.image,
            "created_at": posm.created_at.strftime("%Y-%m-%d %H:%M"),  # cleaner format
        })


    return render(
        request,
        "restaurants.html",
        {
            "restaurants": page_obj,
            "username": username,
            "page_obj": page_obj,
            "query": query, 
            "cooler_list":cooler_list,
            "posm_list":posm_list
        },
    )

from openpyxl.styles import Protection

def download_template(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return HttpResponse("User not logged in", status=401)
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return HttpResponse("User not found", status=404)

    if user.is_superuser:
        restaurants = Restaurant.objects.all()
    else:
        restaurants = Restaurant.objects.none()
        if hasattr(user, "gm_profile") and user.gm_profile:
            restaurants = restaurants | user.gm_profile.restaurants.all()
        if hasattr(user, "rsm_profile") and user.rsm_profile:
            restaurants = restaurants | user.rsm_profile.restaurants.all()
        restaurants = restaurants.distinct()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cases Sales Template"

    ws.column_dimensions['A'].width = 25  
    ws.column_dimensions['B'].width = 40  
    ws.column_dimensions['C'].width = 15  
    ws.column_dimensions['D'].width = 15  
    ws.column_dimensions['E'].width = 20  

    headers = ["Restaurant Code", "Restaurant Name", "Month Number", "Week Number", "Cases Sales"]
    ws.append(headers)

    for restaurant in restaurants:
        row = [
            restaurant.restaurant_code,
            restaurant.restaurant_name,
            "",  
            "",  
            ""  
        ]
        ws.append(row)

    ws.protection.sheet = True

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row[2:5]:
            cell.protection = Protection(locked=False)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename="cases_sales_template.xlsx"'
    wb.save(response)
    return response


from openpyxl import load_workbook, Workbook


from openpyxl import load_workbook


def month_number_to_name(month_num):
    """Convert month number (1-12) to full month name"""
    try:
        month_num = int(month_num)
        return calendar.month_name[month_num] 
    except:
        return None

def upload_crate_sales(request):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    if "file" not in request.FILES:
        return JsonResponse({"error": "No file uploaded"}, status=400)

    excel_file = request.FILES["file"]

    try:
        wb = load_workbook(filename=excel_file)
        ws = wb.active
    except Exception as e:
        return JsonResponse({"error": f"Invalid Excel file: {str(e)}"}, status=400)

    rows_created = 0
    skipped_rows = 0
    errors = []

    current_year = datetime.now().year 

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        restaurant_code = row[0]  
        month_value = row[2]      
        week_number = row[3]      
        cases_sales = row[4]       

        if not (month_value and week_number and cases_sales):
            skipped_rows += 1
            continue

        try:
            restaurant = Restaurant.objects.get(restaurant_code=restaurant_code)
        except Restaurant.DoesNotExist:
            errors.append(f"Row {idx}: Restaurant code {restaurant_code} not found")
            continue

        if isinstance(month_value, (int, float)):
            month_name = month_number_to_name(month_value)
        else:
            month_name = str(month_value).strip()

        sale_month = f"{month_name} {current_year}"

        Restaurant_Crate_Sales.objects.create(
            sale_month=sale_month,
            saleweek=str(week_number).strip(),
            crate_quantity=int(cases_sales),
            restaurant=restaurant,
            is_approved=True
        )
        rows_created += 1

    return JsonResponse({
        "message": f"Successfully uploaded {rows_created} restaurants cases sales",
        "skipped_rows": skipped_rows,
        "errors": errors
    })



def create_restaurant_api(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Only POST allowed"}, status=405)

    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid token: {e}"}, status=401)

    if not (user.is_superuser or user.is_staff):
        return JsonResponse({"success": False, "message": "Permission denied"}, status=403)

    try:
        data = json.loads(request.body.decode("utf-8"))

        name = data.get("name")
        manager_id = data.get("manager")  
        contact_no = data.get("contactNo")
        category = data.get("category")

        if not name or not manager_id:
            return JsonResponse({"success": False, "message": "Name and Manager are required"}, status=400)

        try:
            manager = User.objects.get(id=manager_id)
        except User.DoesNotExist:
            return JsonResponse({"success": False, "message": "Invalid manager ID"}, status=400)

        if Restaurant.objects.filter(manager=manager).exists():
            return JsonResponse({
            "success": False,
            "message": "This manager is already assigned to a restaurant"
        }, status=400)
     
        restaurant = Restaurant.objects.create(
            restaurant_name=name,
            
            manager=manager,
            manager_phone_number=contact_no,
            restaurant_category=category,
        )

        return JsonResponse({
            "success": True,
            "message": "Restaurant created successfully",
        }, status=201)

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON format"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)
    

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from rest_framework_simplejwt.tokens import AccessToken
from django.contrib.auth.models import User
import json

def create_restaurant_target_api(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Only POST allowed"}, status=405)

    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)
    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid token: {e}"}, status=401)

    if not (user.is_superuser or user.is_staff):
        return JsonResponse({"success": False, "message": "Permission denied"}, status=403)

    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON format"}, status=400)

    target_type = data.get("target_type", "").strip()
    target_month = data.get("month")
    target_value = data.get("target_value")
    restaurant_id = data.get("restaurant_id")

    print(restaurant_id)

    if not target_month or target_value is None:
        return JsonResponse({"success": False, "message": "Month and target_value are required"}, status=400)

    if not restaurant_id:
        return JsonResponse({"success": False, "message": "restaurant_id is required"}, status=400)

    try:
        target_value = int(target_value)  # or float if decimals allowed
    except (TypeError, ValueError):
        return JsonResponse({"success": False, "message": "target_value must be a number."}, status=400)

    try:
        restaurant_pk = int(restaurant_id)
    except (TypeError, ValueError):
        return JsonResponse({"success": False, "message": "restaurant_id must be an integer"}, status=400)

    try:
        restaurant_obj = Restaurant.objects.get(pk=restaurant_pk)
    except Restaurant.DoesNotExist:
        return JsonResponse({"success": False, "message": "Restaurant not found"}, status=404)

    import datetime
    current_month = calendar.month_abbr[datetime.date.today().month]  # e.g. "September"
    if target_month.lower() != current_month.lower():
        return JsonResponse(
            {"success": False, "message": f"Targets can only be created for the current month ({current_month})."},
            status=400
        )

    if target_type in ["Compliance", "Cases Purchased", "Consumer"]:
        exists = Restaurant_Target.objects.filter(
            restaurant=restaurant_obj,
            target_type=target_type,
            target_month=target_month
        ).exists()
        if exists:
            return JsonResponse(
                {"success": False, "message": f"Target for {target_type} already exists for {target_month}."},
                status=400
            )

    if target_type == "Compliance" and (target_value < 0 or target_value > 100):
        return JsonResponse(
            {"success": False, "message": "Compliance target must be between 0 and 100 (percentage)."},
            status=400
        )

    rt = Restaurant_Target.objects.create(
        restaurant=restaurant_obj,
        target_type=target_type,
        target_month=target_month,
        target_value=target_value,
        target_status=True,
    )

    return JsonResponse({
        "success": True,
        "message": "Restaurant Target created successfully",
        "id": rt.id,
    }, status=200)


def edit_restaurant_target_api(request, target_id):
    if request.method != "PUT":
        return JsonResponse({"success": False, "message": "Only PUT allowed"}, status=405)

    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid token: {e}"}, status=401)

    if not (user.is_superuser or user.is_staff):
        return JsonResponse({"success": False, "message": "Permission denied"}, status=403)

    # --- Parse body ---
    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON"}, status=400)

    print(data)

    target_type = data.get("target_type", "").strip()
    target_month = data.get("target_month")
    target_value = data.get("target_value")
    target_status = data.get("target_status", "Inactive")



    try:
        target_value = int(target_value)  # or float(target_value) if decimals are allowed
    except (TypeError, ValueError):
        return JsonResponse(
            {"success": False, "message": "Target value must be a number."},
            status=400
        )

    import datetime
    current_month = calendar.month_abbr[datetime.date.today().month]  # e.g., "September"
    if target_month.lower() != current_month.lower():
        return JsonResponse(
            {"success": False, "message": f"Targets can only be edited for the current month ({current_month})."},
            status=400
        )
    

    if target_type == "Compliance":
        if target_value < 0 or target_value > 100:
            return JsonResponse(
                {"success": False, "message": "Compliance target must be between 0 and 100 (percentage)."},
                status=400
            )



    if not target_month or target_value is None:
        return JsonResponse({"success": False, "message": "Month and target_value are required"}, status=400)

    try:
        target_value = int(target_value)
    except (TypeError, ValueError):
        return JsonResponse({"success": False, "message": "target_value must be an integer"}, status=400)

    try:
        rt = Restaurant_Target.objects.get(pk=target_id)
    except Restaurant_Target.DoesNotExist:
        return JsonResponse({"success": False, "message": "Target not found"}, status=404)

    # --- Update fields ---
    rt.target_type = target_type
    rt.target_month = target_month
    rt.target_value = target_value
    rt.target_status = True if target_status == "Active" else False
    rt.save()

    return JsonResponse({
        "success": True,
        "message": "Restaurant Target updated successfully",
        "id": rt.id,
    }, status=200)


# @csrf_exempt
# def edit_restaurant_api(request, restaurant_id):
#     if request.method != "PUT":
#         return JsonResponse({"success": False, "message": "Only PUT allowed"}, status=405)

#     auth_header = request.headers.get("Authorization")
#     if not auth_header or not auth_header.startswith("Bearer "):
#         return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

#     token_str = auth_header.split(" ")[1]
#     try:
#         access_token = AccessToken(token_str)
#         user_id = access_token["user_id"]
#         user = User.objects.get(id=user_id)
#     except Exception as e:
#         return JsonResponse({"success": False, "message": f"Invalid or expired token: {e}"}, status=401)

#     if not (user.is_superuser or user.is_staff):
#         return JsonResponse({"success": False, "message": "Permission denied"}, status=403)

#     try:
#         restaurant = Restaurant.objects.get(id=restaurant_id)
#     except Restaurant.DoesNotExist:
#         return JsonResponse({"success": False, "message": "Restaurant not found"}, status=404)

#     try:
#         data = json.loads(request.body.decode("utf-8"))

#         name = data.get("name")
#         manager_id = data.get("manager")  
#         contact_no = data.get("contactNo")

#         if not name:
#             return JsonResponse({"success": False, "message": "Restaurant name is required"}, status=400)

#         manager = None
#         if manager_id:
#             try:
#                 manager = User.objects.get(id=manager_id)
#                 if Restaurant.objects.filter(manager=manager).exclude(id=restaurant.id).exists():
#                     return JsonResponse({"success": False, "message": "Manager is already assigned to another restaurant"}, status=400)
#             except User.DoesNotExist:
#                 return JsonResponse({"success": False, "message": "Invalid manager ID"}, status=400)

#         restaurant.restaurant_name = name
#         restaurant.manager = manager if manager else restaurant.manager
#         restaurant.manager_phone_number = contact_no
#         restaurant.save()

#         return JsonResponse({
#             "success": True,
#             "message": "Restaurant updated successfully",
#             "restaurant_id": restaurant.id,
#             "restaurant_code": restaurant.restaurant_code,
#         }, status=200)

#     except json.JSONDecodeError:
#         return JsonResponse({"success": False, "message": "Invalid JSON format"}, status=400)
#     except Exception as e:
#         return JsonResponse({"success": False, "message": str(e)}, status=500)


@csrf_exempt
def delete_restaurant_api(request, restaurant_id):
    if request.method != "DELETE":
        return JsonResponse({"success": False, "message": "Only DELETE allowed"}, status=405)

    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid or expired token: {e}"}, status=401)

    if not (user.is_superuser or user.is_staff):
        return JsonResponse({"success": False, "message": "Permission denied"}, status=403)

    try:
        restaurant = Restaurant.objects.get(id=restaurant_id)
        restaurant_name = restaurant.restaurant_name
        restaurant.delete()

        return JsonResponse({
            "success": True,
            "message": f"Restaurant '{restaurant_name}' deleted successfully"
        }, status=200)

    except Restaurant.DoesNotExist:
        return JsonResponse({"success": False, "message": "Restaurant not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)


@login_required(login_url='/dashboard/login/')
def details_restaurant(request, restaurant_id):
    user = request.user
    username = user.username
    restaurant = Restaurant.objects.get(id=restaurant_id)

    selected_bottler_name = restaurant.bottler.id
    print(selected_bottler_name)
    selected_bottler_obj = Bottler.objects.filter(id=selected_bottler_name).first()

    if user.is_superuser:
        all_bottlers = Bottler.objects.all()
    else:
        all_bottlers = None

    all_rsms = RSM.objects.filter(bottler=selected_bottler_obj) if selected_bottler_obj else RSM.objects.none()
    all_gms = GM.objects.filter(bottler=selected_bottler_obj) if selected_bottler_obj else GM.objects.none()

    restaurant_target = Restaurant_Target.objects.filter(restaurant=restaurant)

    bottle_crate_target = Restaurant_Crate_Sales.objects.filter(restaurant=restaurant).order_by("-created_at")


    bottle_paginator = Paginator(bottle_crate_target, 10)  # show 10 per page
    bottle_page_number = request.GET.get("page")
    page_obj_bottle = bottle_paginator.get_page(bottle_page_number)

    restaurant_paginator = Paginator(restaurant_target, 10)  # show 10 per page
    restaurant_page_number = request.GET.get("page")
    page_obj_restaurant = restaurant_paginator.get_page(restaurant_page_number)

    return render(
        request,
        "details-restaurants.html",
        {
            "user": user,
            "username":username,
            "restaurant": restaurant,
            "all_rsms": all_rsms,
            "all_gms": all_gms,
            "all_bottlers": all_bottlers,
            "selected_bottler_obj": selected_bottler_obj,
            "selected_bottler_name": selected_bottler_name,
            "restaurant_target":page_obj_restaurant,
            "bottle_crate_target":page_obj_bottle,
        }
    )

@login_required(login_url='/dashboard/login/')
def get_rsms_gms_by_bottler(request, bottler_id):
    try:
        bottler = Bottler.objects.get(id=bottler_id)
        rsms = list(RSM.objects.filter(bottler=bottler).values("id", "rsm_name"))
        gms = list(GM.objects.filter(bottler=bottler).values("id", "gm_name"))
        return JsonResponse({"success": True, "rsms": rsms, "gms": gms})
    except Bottler.DoesNotExist:
        return JsonResponse({"success": False, "rsms": [], "gms": []})

@login_required(login_url='/dashboard/login/')
def get_restaurant_target(request, target_id):
    if request.method != "GET":
        return JsonResponse({"success": False, "message": "Only DELETE allowed"}, status=405)

    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid or expired token: {e}"}, status=401)

    if not (user.is_superuser or user.is_staff):
        return JsonResponse({"success": False, "message": "Permission denied"}, status=403)


    try:
        target = Restaurant_Target.objects.get(id=target_id)
        return JsonResponse({
            "success": True,
            "target": {
                "id": target.id,
                "target_type": target.target_type,        
                "target_month": target.target_month,     
                "target_value": target.target_value, 
                "target_status": "Active" if target.target_status else "Inactive"
            }
        })
    except Restaurant_Target.DoesNotExist:
        return JsonResponse({"success": False, "message": "Target not found"}, status=404)

@login_required(login_url='/dashboard/login/')
def get_crate_target(request, crate_id):
    if request.method != "GET":
        return JsonResponse({"success": False, "message": "Only DELETE allowed"}, status=405)

    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid or expired token: {e}"}, status=401)


    try:
        crate = Restaurant_Crate_Sales.objects.get(id=crate_id)
        return JsonResponse({
            "success": True,
            "crate": {
                "id": crate.id,
                "requested_at": crate.created_at,        
                "is_approved": "Approved" if crate.is_approved else ("Declined" if crate.is_declined else "Requested")
            }
        })
    except Restaurant_Target.DoesNotExist:
        return JsonResponse({"success": False, "message": "Target not found"}, status=404)
    



def edit_crate_target_api(request, crate_id):
    if request.method != "PUT":
        return JsonResponse({"success": False, "message": "Only PUT allowed"}, status=405)

    # --- JWT Auth ---
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid token: {e}"}, status=401)



    # --- Parse body ---
    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON"}, status=400)

    print(data)

    is_approved = data.get("is_approved")
 

    try:
        crate = Restaurant_Crate_Sales.objects.get(id=crate_id)
    except Restaurant_Crate_Sales.DoesNotExist:
        return JsonResponse({"success": False, "message": "Crate Target not found"}, status=404)

    # --- Update fields ---
    if is_approved == "Approved":
        crate.is_approved = True
    else:
        crate.is_approved = False
        crate.is_declined = True

    crate.save()

    return JsonResponse({
        "success": True,
        "message": "Crate Target updated successfully",
        "id": crate.id,
    }, status=200)


def get_cooler(request, cooler_id):
    if request.method != "GET":
        return JsonResponse({"success": False, "message": "Only GET allowed"}, status=405)

    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid or expired token: {e}"}, status=401)


    try:
        cooler = Restaurant_Cooler.objects.get(id=cooler_id)
        return JsonResponse({
            "success": True,
            "cooler": {
                "id": cooler_id,
                "requested_at": cooler.created_at,        
                "raw_data":cooler.raw_data,
            }
        })
    except Restaurant_Cooler.DoesNotExist:
        return JsonResponse({"success": False, "message": "cooler not found"}, status=404)
    

def edit_cooler_api(request, cooler_id):
    if request.method != "PUT":
        return JsonResponse({"success": False, "message": "Only PUT allowed"}, status=405)

    # --- JWT Auth ---
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid token: {e}"}, status=401)



    # --- Parse body ---
    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON"}, status=400)

    print(data)

    raw_data = data.get("raw_data")
 

    try:
        cooler = Restaurant_Cooler.objects.get(id=cooler_id)
    except Restaurant_Cooler.DoesNotExist:
        return JsonResponse({"success": False, "message": "Cooler not found"}, status=404)

    # --- Update fields ---
    cooler.is_manually = False
    cooler.is_changed = True
    cooler.raw_data = raw_data
    cooler.save()

    return JsonResponse({
        "success": True,
        "message": "Cooler updated successfully",
        "id": cooler.id,
    }, status=200)



def get_posm(request, posm_id):
    if request.method != "GET":
        return JsonResponse({"success": False, "message": "Only GET allowed"}, status=405)

    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid or expired token: {e}"}, status=401)


    try:
        posm = Restaurant_POSM.objects.get(id=posm_id)

        return JsonResponse({
            "success": True,
            "posm": {
                "id": posm_id,
                "requested_at": posm.created_at,        
                "raw_data":posm.raw_data,
            }
        })
    except Restaurant_POSM.DoesNotExist:
        return JsonResponse({"success": False, "message": "posm not found"}, status=404)
    

def edit_posm_api(request, posm_id):
    if request.method != "PUT":
        return JsonResponse({"success": False, "message": "Only PUT allowed"}, status=405)

    # --- JWT Auth ---
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid token: {e}"}, status=401)



    # --- Parse body ---
    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON"}, status=400)

    print(data)

    raw_data = data.get("raw_data")
 

    try:
        posm = Restaurant_POSM.objects.get(id=posm_id)
    except Restaurant_POSM.DoesNotExist:
        return JsonResponse({"success": False, "message": "Posm not found"}, status=404)

    # --- Update fields ---
    posm.is_manually = False
    posm.is_changed = True
    posm.raw_data = raw_data
    posm.save()

    return JsonResponse({
        "success": True,
        "message": "Posm updated successfully",
        "id": posm.id,
    }, status=200)


@developer_login_required
def edit_developer_cooler_api(request, cooler_id):
    if request.method != "PUT":
        return JsonResponse({"success": False, "message": "Only PUT allowed"}, status=405)

    # --- JWT Auth ---
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid token: {e}"}, status=401)



    # --- Parse body ---
    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON"}, status=400)


    cooler_status = data.get("cooler_status")
 

    try:
        cooler = Restaurant_Cooler.objects.get(id=cooler_id)
    except Restaurant_Cooler.DoesNotExist:
        return JsonResponse({"success": False, "message": "Cooler not found"}, status=404)

    print("cooler_status",cooler_status)

    # --- Update fields ---
    if cooler_status == "Approved":
        cooler.is_checked = True
    else:
        cooler.is_checked = False
    cooler.save()

    return JsonResponse({
        "success": True,
        "message": "Cooler updated successfully",
        "id": cooler.id,
    }, status=200)


@developer_login_required
def edit_developer_posm_api(request, posm_id):
    if request.method != "PUT":
        return JsonResponse({"success": False, "message": "Only PUT allowed"}, status=405)

    # --- JWT Auth ---
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return JsonResponse({"success": False, "message": "Missing or invalid token"}, status=401)

    token_str = auth_header.split(" ")[1]
    try:
        access_token = AccessToken(token_str)
        user_id = access_token["user_id"]
        user = User.objects.get(id=user_id)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Invalid token: {e}"}, status=401)



    # --- Parse body ---
    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON"}, status=400)


    posm_status = data.get("posm_status")
 

    try:
        posm = Restaurant_POSM.objects.get(id=posm_id)
    except Restaurant_POSM.DoesNotExist:
        return JsonResponse({"success": False, "message": "Cooler not found"}, status=404)

    print("posm_status",posm_status)

    # --- Update fields ---
    if posm_status == "Approved":
        posm.is_checked = True
    else:
        posm.is_checked = False
    posm.save()

    return JsonResponse({
        "success": True,
        "message": "Posm updated successfully",
        "id": posm.id,
    }, status=200)




@login_required(login_url='/dashboard/login/')

@superuser_required
def lucky_draw(request):
    user_id = request.session.get("user_id")
    user = User.objects.get(id=user_id)
    username = user.username
    return render(request,"lucky-draw.html",{"username":username})

@login_required(login_url='/dashboard/login/')
@superuser_required
def done_draws(request):
    user_id = request.session.get("user_id")
    user = User.objects.get(id=user_id)
    username = user.username
    return render(request,"done-draws.html",{"username":username})

@login_required(login_url='/dashboard/login/')
def report(request):
    user_id = request.session.get("user_id")
    user = User.objects.get(id=user_id)
    username = user.username

    if user.is_superuser:
        bottlers = Bottler.objects.all()
        restaurants = Restaurant.objects.all()
        gms = GM.objects.all()
        rsms = RSM.objects.all()

    else:
        try:
            gm = GM.objects.get(user_id=user_id)
            gm_restaurants = gm.restaurants.all()
            bottlers = Bottler.objects.filter(id=gm.bottler.id)
            restaurants = Restaurant.objects.filter(id__in=gm_restaurants.values_list("id", flat=True))
            gms = GM.objects.filter(restaurants__in=gm_restaurants).distinct()
            rsms = RSM.objects.filter(restaurants__in=gm_restaurants).distinct()

        except GM.DoesNotExist:
            try:
                rsm = RSM.objects.get(user_id=user_id)
                rsm_restaurants = rsm.restaurants.all()
                bottlers = Bottler.objects.filter(id=rsm.bottler.id)
                restaurants = Restaurant.objects.filter(id__in=rsm_restaurants.values_list("id", flat=True))
                gms = GM.objects.filter(restaurants__in=rsm_restaurants).distinct()
                rsms = RSM.objects.filter(restaurants__in=rsm_restaurants).distinct()
            except RSM.DoesNotExist:
                bottlers = Bottler.objects.none()


    return render(request,"report.html",{"username":username,"bottlers":bottlers})



def generate_report(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)

    try:
        data = json.loads(request.body)

        start_date = data.get("startDate")
        end_date = data.get("endDate")
        bottler_id = data.get("bottler")
        gm_ids = data.get("gms", [])
        rsm_ids = data.get("rsms", [])
        restaurant_ids = data.get("restaurants", [])
        kpi = data.get("kpi")

        filters = {}
        if start_date:
            filters["created_at__gte"] = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            filters["created_at__lte"] = datetime.strptime(end_date, "%Y-%m-%d")

        bottlers = Bottler.objects.all()
        if bottler_id:
            bottlers = bottlers.filter(id=bottler_id)

        results = {"bottlers": []}

        for bottler in bottlers:
            bottler_data = {
                "bottler_name": bottler.bottler_name,
                "gms": [],
                "rsms": []
            }

            # ---------------- GMs ----------------
            gms = bottler.gms.all()
            if gm_ids:
                gms = gms.filter(id__in=gm_ids)

            for gm in gms:
                gm_data = {
                    "gm_name": gm.gm_name,
                    "gm_phone_number": gm.gm_phone_number,
                    "restaurants": []
                }
                restaurants = gm.restaurants.all()
                if restaurant_ids:
                    restaurants = restaurants.filter(id__in=restaurant_ids)

                for restaurant in restaurants:
                    gm_data["restaurants"].append(
                        _get_restaurant_data(restaurant, filters, kpi)
                    )
                bottler_data["gms"].append(gm_data)

            # ---------------- RSMs ----------------
            rsms = bottler.rsm.all()
            if rsm_ids:
                rsms = rsms.filter(id__in=rsm_ids)

            for rsm in rsms:
                rsm_data = {
                    "rsm_name": rsm.rsm_name,
                    "rsm_phone_number": rsm.rsm_phone_number,
                    "restaurants": []
                }
                restaurants = rsm.restaurants.all()
                if restaurant_ids:
                    restaurants = restaurants.filter(id__in=restaurant_ids)

                for restaurant in restaurants:
                    rsm_data["restaurants"].append(
                        _get_restaurant_data(restaurant, filters, kpi)
                    )
                bottler_data["rsms"].append(rsm_data)

            results["bottlers"].append(bottler_data)

        return JsonResponse(results, safe=False)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def _get_restaurant_data(restaurant, filters, kpi):
    """Return restaurant data, filtered by KPI if provided."""

    data = {
        "restaurant_name": restaurant.restaurant_name,
        "restaurant_category": restaurant.restaurant_category,
        "manager": restaurant.manager.name if restaurant.manager else None,
        "tagged_bottler": restaurant.bottler.bottler_name if restaurant.bottler else None,
    }

    if not kpi or kpi == "Cases Purchased":
        sales_achieved_target = Restaurant_Crate_Sales.objects.filter(
            restaurant=restaurant, is_approved=True, **filters
        ).aggregate(total=Sum("crate_quantity"))["total"] or 0

        sales_total_target = Restaurant_Target.objects.filter(
            restaurant=restaurant, target_type="Cases Purchased", **filters
        ).aggregate(total=Sum("target_value"))["total"] or 0

        raw_sales = list(Restaurant_Crate_Sales.objects.filter(
            restaurant=restaurant, **filters
        ).values(
            "sale_month", "crate_quantity", "restaurant__restaurant_name",
            "created_at", "is_approved", "is_declined"
        ))

        data["cases_sales"] = {
            "achieved_target": sales_achieved_target,
            "total_target": sales_total_target,
            "raw_sales": raw_sales,
        }

    if not kpi or kpi == "Compliance":
        data["coolers"] = list(Restaurant_Cooler.objects.filter(
            restaurant=restaurant, **filters
        ).values(
            "cooler_type", "restaurant__restaurant_name",
            "created_at", "raw_data","is_checked"
        ))

    if not kpi or kpi == "Compliance":
        data["posms"] = list(Restaurant_POSM.objects.filter(
            restaurant=restaurant, **filters
        ).values(
            "posm_type", "restaurant__restaurant_name",
            "created_at", "raw_data","is_checked"
        ))

    if not kpi or kpi == "Consumers":
        data["consumers"] = list(Consumer.objects.filter(
            restaurant=restaurant, **filters
        ).values(
            "consumer_name", "consumer_phone_number",
            "restaurant__restaurant_name", "created_at"
        ))

    return data


def _get_restaurant_data(restaurant, filters, kpi):
    """Return restaurant data with KPI filtering."""

    # --- Sales ---
    sales_achieved_target = Restaurant_Crate_Sales.objects.filter(
        restaurant=restaurant, is_approved=True, **filters
    ).aggregate(total=Sum("crate_quantity"))["total"] or 0

    sales_total_target = Restaurant_Target.objects.filter(
        restaurant=restaurant, target_type="Cases Purchased", **filters
    ).aggregate(total=Sum("target_value"))["total"] or 0

    raw_sales = list(Restaurant_Crate_Sales.objects.filter(
        restaurant=restaurant, **filters
    ).values(
        "sale_month", "crate_quantity", "restaurant__restaurant_name",
        "created_at", "is_approved", "is_declined"
    ))

    # --- Coolers ---
    coolers = list(Restaurant_Cooler.objects.filter(
        restaurant=restaurant, **filters
    ).values("cooler_type", "restaurant__restaurant_name",
             "created_at", "raw_data", "is_checked"))

    # --- POSMs ---
    posms = list(Restaurant_POSM.objects.filter(
        restaurant=restaurant, **filters
    ).values("posm_type", "restaurant__restaurant_name",
             "created_at", "raw_data", "is_checked"))

    # --- Consumers ---
    consumer_achieved_target = Consumer.objects.filter(
        restaurant=restaurant, is_checked=True, **filters
    ).count()

    consumer_total_target = Restaurant_Target.objects.filter(
        restaurant=restaurant, target_type="Consumer", **filters
    ).aggregate(total=Sum("target_value"))["total"] or 0

    consumers = list(Consumer.objects.filter(
        restaurant=restaurant, **filters
    ).values("consumer_name", "consumer_phone_number",
             "restaurant__restaurant_name", "created_at"))

    return {
        "restaurant_name": restaurant.restaurant_name,
        "restaurant_category": restaurant.restaurant_category,
        "manager": restaurant.manager.name if restaurant.manager else None,
        "tagged_bottler": restaurant.bottler.bottler_name if restaurant.bottler else None,
        "cases_sales": {
            "achieved_target": sales_achieved_target,
            "total_target": sales_total_target,
            "raw_sales": raw_sales,
        },
        "coolers": coolers,
        "posms": posms,
        "consumers": {
            "achieved_target": consumer_achieved_target,
            "total_target": consumer_total_target,
            "consumers_data": consumers,
        }
    }



import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import io, json
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

def export_report_excel(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)

    try:
        data = json.loads(request.body)

        start_date = data.get("startDate")
        end_date = data.get("endDate")
        bottler_id = data.get("bottler")
        gm_ids = data.get("gms", [])
        rsm_ids = data.get("rsms", [])
        restaurant_ids = data.get("restaurants", [])
        kpi = data.get("kpi")

        filters = {}
        if start_date:
            filters["created_at__gte"] = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            filters["created_at__lte"] = datetime.strptime(end_date, "%Y-%m-%d")

        bottlers = Bottler.objects.all()
        if bottler_id:
            bottlers = bottlers.filter(id=bottler_id)

        # --- Create workbook ---
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # --- Summary header ---
        summary_headers = [
            "Bottler", "GM / RSM", "Restaurant", "Category", "Manager",
            "Sales Achieved", "Sales Target", "Consumer Achieved", "Consumer Target"
        ]
        ws_summary.append(summary_headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        for col in range(1, len(summary_headers) + 1):
            cell = ws_summary.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # --- Extra sheets ---
        ws_sales = wb.create_sheet("Sales")
        ws_coolers = wb.create_sheet("Coolers")
        ws_posms = wb.create_sheet("POSMs")
        ws_consumers = wb.create_sheet("Consumers")

        ws_sales.append(["Bottler", "GM / RSM", "Restaurant", "Sale Month", "Quantity", "Approved", "Declined", "Created At"])
        ws_coolers.append(["Bottler", "GM / RSM", "Restaurant", "Cooler Type", "Checked", "Created At", "Raw Data"])
        ws_posms.append(["Bottler", "GM / RSM", "Restaurant", "POSM Type", "Checked", "Created At", "Raw Data"])
        ws_consumers.append(["Bottler", "GM / RSM", "Restaurant", "Consumer Name", "Phone", "Created At"])

        # --- Populate ---
        for bottler in bottlers:
            # GMs
            gms = bottler.gms.all()
            if gm_ids:
                gms = gms.filter(id__in=gm_ids)

            for gm in gms:
                gm_name = f"GM: {gm.gm_name}"
                restaurants = gm.restaurants.all()
                if restaurant_ids:
                    restaurants = restaurants.filter(id__in=restaurant_ids)

                for restaurant in restaurants:
                    r_data = _get_restaurant_data(restaurant, filters, kpi)

                    # Summary
                    ws_summary.append([
                        bottler.bottler_name,
                        gm_name,
                        r_data["restaurant_name"],
                        r_data["restaurant_category"],
                        r_data["manager"],
                        r_data["cases_sales"]["achieved_target"],
                        r_data["cases_sales"]["total_target"],
                        r_data["consumers"]["achieved_target"],
                        r_data["consumers"]["total_target"],
                    ])

                    # Sales
                    for sale in r_data["cases_sales"]["raw_sales"]:
                        ws_sales.append([
                            bottler.bottler_name,
                            gm_name,
                            sale["restaurant__restaurant_name"],
                            sale["sale_month"],
                            sale["crate_quantity"],
                            sale["is_approved"],
                            sale["is_declined"],
                            sale["created_at"].strftime("%Y-%m-%d %H:%M:%S") if sale["created_at"] else "",
                        ])

                    # Coolers
                    for cooler in r_data["coolers"]:
                        ws_coolers.append([
                            bottler.bottler_name,
                            gm_name,
                            cooler["restaurant__restaurant_name"],
                            cooler["cooler_type"],
                            cooler["is_checked"],
                            cooler["created_at"].strftime("%Y-%m-%d %H:%M:%S") if cooler["created_at"] else "",
                            cooler["raw_data"],
                        ])

                    # POSMs
                    for posm in r_data["posms"]:
                        ws_posms.append([
                            bottler.bottler_name,
                            gm_name,
                            posm["restaurant__restaurant_name"],
                            posm["posm_type"],
                            posm["is_checked"],
                            posm["created_at"].strftime("%Y-%m-%d %H:%M:%S") if posm["created_at"] else "",
                            posm["raw_data"],
                        ])

                    # Consumers
                    for consumer in r_data["consumers"]["consumers_data"]:
                        ws_consumers.append([
                            bottler.bottler_name,
                            gm_name,
                            consumer["restaurant__restaurant_name"],
                            consumer["consumer_name"],
                            consumer["consumer_phone_number"],
                            consumer["created_at"].strftime("%Y-%m-%d %H:%M:%S") if consumer["created_at"] else "",
                        ])

            # RSMs
            rsms = bottler.rsm.all()
            if rsm_ids:
                rsms = rsms.filter(id__in=rsm_ids)

            for rsm in rsms:
                rsm_name = f"RSM: {rsm.rsm_name}"
                restaurants = rsm.restaurants.all()
                if restaurant_ids:
                    restaurants = restaurants.filter(id__in=restaurant_ids)

                for restaurant in restaurants:
                    r_data = _get_restaurant_data(restaurant, filters, kpi)

                    ws_summary.append([
                        bottler.bottler_name,
                        rsm_name,
                        r_data["restaurant_name"],
                        r_data["restaurant_category"],
                        r_data["manager"],
                        r_data["cases_sales"]["achieved_target"],
                        r_data["cases_sales"]["total_target"],
                        r_data["consumers"]["achieved_target"],
                        r_data["consumers"]["total_target"],
                    ])

                    for sale in r_data["cases_sales"]["raw_sales"]:
                        ws_sales.append([
                            bottler.bottler_name,
                            rsm_name,
                            sale["restaurant__restaurant_name"],
                            sale["sale_month"],
                            sale["crate_quantity"],
                            sale["is_approved"],
                            sale["is_declined"],
                            sale["created_at"].strftime("%Y-%m-%d %H:%M:%S") if sale["created_at"] else "",
                        ])

                    for cooler in r_data["coolers"]:
                        ws_coolers.append([
                            bottler.bottler_name,
                            rsm_name,
                            cooler["restaurant__restaurant_name"],
                            cooler["cooler_type"],
                            cooler["is_checked"],
                            cooler["created_at"].strftime("%Y-%m-%d %H:%M:%S") if cooler["created_at"] else "",
                            cooler["raw_data"],
                        ])

                    for posm in r_data["posms"]:
                        ws_posms.append([
                            bottler.bottler_name,
                            rsm_name,
                            posm["restaurant__restaurant_name"],
                            posm["posm_type"],
                            posm["is_checked"],
                            posm["created_at"].strftime("%Y-%m-%d %H:%M:%S") if posm["created_at"] else "",
                            posm["raw_data"],
                        ])

                    for consumer in r_data["consumers"]["consumers_data"]:
                        ws_consumers.append([
                            bottler.bottler_name,
                            rsm_name,
                            consumer["restaurant__restaurant_name"],
                            consumer["consumer_name"],
                            consumer["consumer_phone_number"],
                            consumer["created_at"].strftime("%Y-%m-%d %H:%M:%S") if consumer["created_at"] else "",
                        ])

        # --- Auto column width ---
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 3

        # --- Save to memory stream ---
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="report.xlsx"'
        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)



@login_required
def api_logout(request):
    if request.method == "POST":
        logout(request)
        return JsonResponse({"success": True, "message": "Logged out successfully"})
    return JsonResponse({"success": False, "message": "Invalid request"}, status=400)
